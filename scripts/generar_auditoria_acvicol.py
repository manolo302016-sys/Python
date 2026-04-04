# -*- coding: utf-8 -*-
"""
generar_auditoria_acvicol.py
============================
Auditor de Data Science — MentalPRO
Empresa: ACVICOL
Fecha: 2026-04-04

Genera un libro Excel (.xlsx) con todas las hojas de auditoría del pipeline
ETL para los 3 visualizadores (V1-Riesgo, V2-Gestión, V3-Gerencial).

Estructura del libro:
  00_PORTADA            — Metadatos de auditoría, parámetros sistema, reglas
  01_ETL_StarSchema     — Paso 1: Respuestas limpias ACVICOL (R1, R6, R17)
  02a_Scoring_Paso1     — Paso 2: Codificación texto→número
  02a_Scoring_Paso2     — Paso 3: Inversión de ítems (R3)
  02a_Scoring_Pasos3_8  — Paso 4: Agrupaciones instrumento/dim/dominio/factor
  02b_Baremos_Dim       — Pasos 9-11: Baremos dimensiones (riesgo/protección)
  02b_Baremos_Dom       — Pasos 12-13: Baremos dominios
  02b_Baremos_Factor    — Paso 14-15: Baremos factores (Intra/Extra/Estrés/Individual)
  06_Benchmarking       — Pasos 16-19: Comparativa vs sector y ENCST Colombia
  07_Frecuencias        — Paso 20: Distribución respuestas por pregunta
  07_Top20              — Paso 20b: Top 20 preguntas vs ENCST
  08_Consolidado        — Paso 21: Fact consolidado con demografía
  V2_Estandarizado      — V2 Sub-paso 1: Estandarización 0-1
  V2_Invertido          — V2 Sub-paso 2: Inversión V2 específica
  V2_Indicadores        — V2 Paso 3: Scores indicadores
  V2_Lineas             — V2 Paso 4: Scores líneas de gestión
  V2_Ejes               — V2 Paso 5: Scores ejes de gestión
  V2_Niveles_Indicadores — V2 Categorización indicadores
  V2_Niveles_Lineas     — V2 Categorización líneas
  V2_Niveles_Ejes       — V2 Categorización ejes
  V2_Protocolos         — V2 Paso 6: Protocolos priorizados
  V2_Vigilancia         — V2 Paso 6b: Vigilancia epidemiológica
  V3_KPIs_Globales      — V3 Paso 1: 19 KPIs ejecutivos
  V3_Demografia         — V3 Paso 2: Distribución demográfica
  V3_Costos             — V3 Paso 3: Cálculo económico (R10)
  V3_Benchmarking       — V3 Paso 4: Benchmarking ejecutivo
  V3_Ranking_Areas      — V3 Paso 5: Ranking áreas críticas
  ZZ_RESUMEN            — Resumen ejecutivo consolidado
  ZZ_ALERTAS            — Alertas y puntos de chequeo para auditoría
"""

from pathlib import Path
from datetime import datetime

import pandas as pd
import numpy as np
import yaml
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Rutas ──────────────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parents[1]
PROC = ROOT / "data" / "processed"
OUT  = ROOT / "output"
OUT.mkdir(parents=True, exist_ok=True)

EMPRESA = "ACVICOL"
FECHA_AUDITORIA = "2026-04-04"

# ── Cargar configuración ──────────────────────────────────────────────────────
with open(ROOT / "config" / "config.yaml", "r", encoding="utf-8") as f:
    CFG = yaml.safe_load(f)

# ── Paleta de colores ─────────────────────────────────────────────────────────
NAVY   = "0A1628"
GOLD   = "C9952A"
CYAN   = "00C2CB"
WHITE  = "FFFFFF"
GRAY_L = "F3F4F6"
GRAY_M = "E5E7EB"

COLOR_RIESGO = {
    1: "10B981",  # sin riesgo - verde
    2: "6EE7B7",  # bajo       - verde claro
    3: "F59E0B",  # medio      - amarillo
    4: "F97316",  # alto       - naranja
    5: "EF4444",  # muy alto   - rojo
}
COLOR_SEMAFORO = {
    "verde":     "10B981",
    "amarillo":  "F59E0B",
    "rojo":      "EF4444",
    "Confidencial": "9CA3AF",
}
COLOR_NIVEL_GESTION = {
    "Gestión prorrogable":        "10B981",
    "Gestión preventiva":         "6EE7B7",
    "Gestión de mejora selectiva":"F59E0B",
    "Intervención correctiva":    "F97316",
    "Intervención urgente":       "EF4444",
}


# ══════════════════════════════════════════════════════════════════════════════
# Helpers de estilo
# ══════════════════════════════════════════════════════════════════════════════

def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def bold_font(color="000000", size=10) -> Font:
    return Font(bold=True, color=color, size=size)

def std_font(color="000000", size=9) -> Font:
    return Font(bold=False, color=color, size=size)

def center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left_align() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def thin_border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def style_header_row(ws, row_num: int, n_cols: int,
                     bg=NAVY, fg=WHITE, height=22):
    """Aplica estilo de encabezado (fila completa)."""
    ws.row_dimensions[row_num].height = height
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill(bg)
        cell.font = bold_font(fg, 9)
        cell.alignment = center()
        cell.border = thin_border()

def style_data_row(ws, row_num: int, n_cols: int, bg=WHITE):
    ws.row_dimensions[row_num].height = 15
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill(bg)
        cell.font = std_font()
        cell.alignment = left_align()
        cell.border = thin_border()

def write_title_block(ws, title: str, subtitle: str = "", empresa: str = EMPRESA):
    """Escribe bloque de título (filas 1-3) en la hoja."""
    ws.merge_cells("A1:Z1")
    t = ws["A1"]
    t.value = title
    t.fill = fill(NAVY)
    t.font = bold_font(WHITE, 14)
    t.alignment = center()
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:Z2")
    s = ws["A2"]
    s.value = f"Empresa: {empresa}  |  {subtitle}  |  Auditoría: {FECHA_AUDITORIA}"
    s.fill = fill(GOLD)
    s.font = bold_font(WHITE, 10)
    s.alignment = center()
    ws.row_dimensions[2].height = 18

def autofit_cols(ws, max_width=50):
    """Ajusta el ancho de columnas según contenido."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), max_width)

def write_df_to_ws(ws, df: pd.DataFrame, start_row: int = 4,
                   header_bg: str = NAVY, col_color_fn=None):
    """Escribe DataFrame a hoja con encabezados estilizados."""
    if df.empty:
        ws.cell(row=start_row, column=1).value = "Sin datos para esta empresa"
        return

    cols = list(df.columns)
    # Encabezados
    for ci, col_name in enumerate(cols, 1):
        cell = ws.cell(row=start_row, column=ci)
        cell.value = col_name
        cell.fill = fill(header_bg)
        cell.font = bold_font(WHITE, 9)
        cell.alignment = center()
        cell.border = thin_border()
    ws.row_dimensions[start_row].height = 22

    # Datos
    for ri, row_data in enumerate(df.itertuples(index=False), start_row + 1):
        bg = GRAY_L if (ri - start_row) % 2 == 0 else WHITE
        ws.row_dimensions[ri].height = 14
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci)
            # Formatear NaN/None
            if val is None or (isinstance(val, float) and np.isnan(val)):
                cell.value = ""
            else:
                cell.value = val
            cell.fill = fill(bg)
            cell.font = std_font(size=8)
            cell.alignment = left_align()
            cell.border = thin_border()
            # Color específico por columna si se pasa función
            if col_color_fn:
                hex_c = col_color_fn(cols[ci - 1], val)
                if hex_c:
                    cell.fill = fill(hex_c)

    autofit_cols(ws)

def color_by_nivel_riesgo(col_name: str, val):
    """Devuelve color de fondo según nivel de riesgo (1-5)."""
    if col_name in ("nivel_riesgo", "nivel") and isinstance(val, (int, float)):
        return COLOR_RIESGO.get(int(val))
    if col_name in ("etiqueta_nivel", "nivel_etiqueta", "etiqueta"):
        etiq_map = {
            "Sin riesgo": "10B981",
            "Riesgo bajo": "6EE7B7",
            "Riesgo medio": "F59E0B",
            "Riesgo alto": "F97316",
            "Riesgo muy alto": "EF4444",
        }
        return etiq_map.get(str(val) if val else "")
    if col_name == "semaforo":
        return COLOR_SEMAFORO.get(str(val) if val else "")
    return None

def color_gestion(col_name: str, val):
    """Color para niveles de gestión V2."""
    if col_name in ("nivel_gestion", "categoria"):
        return COLOR_NIVEL_GESTION.get(str(val) if val else "")
    return color_by_nivel_riesgo(col_name, val)


# ══════════════════════════════════════════════════════════════════════════════
# Carga de datos filtrados para ACVICOL
# ══════════════════════════════════════════════════════════════════════════════

def load_parquet(name: str) -> pd.DataFrame:
    p = PROC / f"{name}.parquet"
    if not p.exists():
        print(f"  [WARN] No existe: {p}")
        return pd.DataFrame()
    df = pd.read_parquet(p)
    return df

def filter_acvicol(df: pd.DataFrame, col: str = "empresa") -> pd.DataFrame:
    if df.empty or col not in df.columns:
        return df
    return df[df[col].str.upper().str.strip() == EMPRESA].copy()

def get_cedulas_acvicol(dim_trabajador: pd.DataFrame) -> set:
    """Devuelve el set de cédulas de ACVICOL."""
    df = filter_acvicol(dim_trabajador)
    return set(df["cedula"].astype(str).unique())

def filter_by_cedula(df: pd.DataFrame, cedulas: set,
                     col: str = "cedula") -> pd.DataFrame:
    if df.empty or col not in df.columns:
        return df
    return df[df[col].astype(str).isin(cedulas)].copy()


# ══════════════════════════════════════════════════════════════════════════════
# Carga global de todos los parquets
# ══════════════════════════════════════════════════════════════════════════════
print(f"\n{'='*60}")
print(f"  AUDITORÍA MentalPRO — {EMPRESA}")
print(f"  Fecha: {FECHA_AUDITORIA}")
print(f"{'='*60}\n")

print("Cargando parquets...")
dim_trab   = load_parquet("dim_trabajador")
dim_demog  = load_parquet("dim_demografia")
dim_aus    = load_parquet("dim_ausentismo")
dim_preg   = load_parquet("dim_pregunta")
cat_anal   = load_parquet("categorias_analisis")

fact_resp  = load_parquet("fact_respuestas_clean")
fact_bruto = load_parquet("fact_scores_brutos")
fact_bar   = load_parquet("fact_scores_baremo")
fact_bench = load_parquet("fact_benchmark")
fact_freq  = load_parquet("fact_frecuencias")
fact_top20 = load_parquet("fact_top20_comparables")
fact_cons  = load_parquet("fact_consolidado")
fact_riesg = load_parquet("fact_riesgo_empresa")

g_estand   = load_parquet("fact_gestion_01_estandarizado")
g_invert   = load_parquet("fact_gestion_02_invertido")
g_indic    = load_parquet("fact_gestion_indicadores")
g_lineas   = load_parquet("fact_gestion_lineas")
g_ejes     = load_parquet("fact_gestion_ejes")
g_niv_ind  = load_parquet("fact_gestion_04_niveles_indicadores")
g_niv_lin  = load_parquet("fact_gestion_04_niveles_lineas")
g_niv_eje  = load_parquet("fact_gestion_04_niveles_ejes")
g_prot     = load_parquet("fact_gestion_05_prioridades_protocolos")
g_vig_res  = load_parquet("fact_gestion_06_vigilancia_resumen")
g_vig_trab = load_parquet("fact_gestion_06_vigilancia_trabajadores")

v3_kpis    = load_parquet("fact_v3_kpis_globales")
v3_demog   = load_parquet("fact_v3_demografia")
v3_costos  = load_parquet("fact_v3_costos")
v3_bench   = load_parquet("fact_v3_benchmarking")
v3_rank    = load_parquet("fact_v3_ranking_areas")

# Obtener cédulas de ACVICOL
CEDULAS = get_cedulas_acvicol(dim_trab)
print(f"  Trabajadores ACVICOL encontrados: {len(CEDULAS)}")

# Filtrar todos los frames por empresa/cédulas
trab_a    = filter_acvicol(dim_trab)
demog_a   = filter_by_cedula(dim_demog, CEDULAS)
aus_a     = filter_by_cedula(dim_aus, CEDULAS)
resp_a    = filter_acvicol(fact_resp)
bruto_a   = filter_by_cedula(fact_bruto, CEDULAS)
bar_a     = filter_acvicol(fact_bar)
bench_a   = filter_acvicol(fact_bench)
freq_a    = filter_acvicol(fact_freq)
top20_a   = filter_acvicol(fact_top20)
cons_a    = filter_acvicol(fact_cons)
riesgo_a  = filter_acvicol(fact_riesg)

g_estand_a  = filter_by_cedula(g_estand, CEDULAS)
g_invert_a  = filter_by_cedula(g_invert, CEDULAS)
g_indic_a   = filter_by_cedula(g_indic, CEDULAS)
g_lineas_a  = filter_by_cedula(g_lineas, CEDULAS)
g_ejes_a    = filter_by_cedula(g_ejes, CEDULAS)
g_niv_ind_a = filter_acvicol(g_niv_ind) if "empresa" in g_niv_ind.columns else g_niv_ind
g_niv_lin_a = filter_acvicol(g_niv_lin) if "empresa" in g_niv_lin.columns else g_niv_lin
g_niv_eje_a = filter_acvicol(g_niv_eje) if "empresa" in g_niv_eje.columns else g_niv_eje
g_prot_a    = filter_acvicol(g_prot) if "empresa" in g_prot.columns else g_prot
g_vig_res_a = filter_acvicol(g_vig_res) if "empresa" in g_vig_res.columns else g_vig_res
g_vig_tr_a  = filter_by_cedula(g_vig_trab, CEDULAS) if "cedula" in g_vig_trab.columns else g_vig_trab

v3_kpis_a   = filter_acvicol(v3_kpis)
v3_demog_a  = filter_acvicol(v3_demog)
v3_cost_a   = filter_acvicol(v3_costos)
v3_bench_a  = filter_acvicol(v3_bench)
v3_rank_a   = filter_acvicol(v3_rank)

print("  Filtrado completado.\n")


# ══════════════════════════════════════════════════════════════════════════════
# CÁLCULO DE MÉTRICAS PARA RESUMEN Y ALERTAS
# ══════════════════════════════════════════════════════════════════════════════

alertas = []  # Lista de (nivel, codigo, descripcion, valor_encontrado, umbral, accion)

def registrar_alerta(nivel: str, codigo: str, descripcion: str,
                     valor="", umbral="", accion=""):
    alertas.append({
        "Nivel": nivel,         # CRITICO / ADVERTENCIA / INFO
        "Código": codigo,
        "Descripción": descripcion,
        "Valor encontrado": str(valor),
        "Umbral / Referencia": str(umbral),
        "Acción recomendada": accion,
    })


# -- Métricas básicas ----------------------------------------------------------
n_trab = len(CEDULAS)
n_forma_a = len(trab_a[trab_a["forma_intra"] == "A"]) if not trab_a.empty else 0
n_forma_b = len(trab_a[trab_a["forma_intra"] == "B"]) if not trab_a.empty else 0
n_respuestas = len(resp_a)
sector = trab_a["sector_rag"].iloc[0] if (not trab_a.empty and "sector_rag" in trab_a.columns) else "No disponible"
sector_raw = trab_a["sector_economico"].iloc[0] if (not trab_a.empty and "sector_economico" in trab_a.columns) else "No disponible"

# -- Alerta: Tamaño de muestra -------------------------------------------------
if n_trab < 5:
    registrar_alerta("CRITICO", "R8-CONF",
                     "Empresa con menos de 5 trabajadores — Regla Confidencialidad R8",
                     n_trab, "≥5", "No reportar resultados individuales ni grupales detallados")
elif n_trab < 20:
    registrar_alerta("ADVERTENCIA", "R8-MUESTRA",
                     "Muestra pequeña (<20 trabajadores) — Interpretación con cautela",
                     n_trab, "≥20", "Señalar limitación estadística en el informe")
else:
    registrar_alerta("INFO", "R8-OK",
                     "Tamaño de muestra aceptable para análisis", n_trab, "≥20", "Sin acción")

# -- Alerta: Sector homologado -------------------------------------------------
if sector == "No clasificado":
    registrar_alerta("ADVERTENCIA", "R17-SECTOR",
                     "Sector económico no pudo homologarse (sector_rag='No clasificado')",
                     sector_raw, "Sector válido ENCST", "Revisar valor en dim_trabajador.sector_economico y actualizar SECTOR_MAP en script 01")
else:
    registrar_alerta("INFO", "R17-OK",
                     f"Sector homologado correctamente: '{sector_raw}' → '{sector}'",
                     sector, "Sector válido ENCST", "Sin acción")

# -- Alerta: Benchmarking global -----------------------------------------------
bench_sector = CFG.get("benchmark_sector", {})
ref_sector = bench_sector.get(sector, bench_sector.get("No clasificado", 39.69))
semaforo_rojo = CFG.get("semaforo_rojo_pct", 35.0)

if not riesgo_a.empty and "pct_alto_muyalto" in riesgo_a.columns:
    pct_emp = riesgo_a["pct_alto_muyalto"].iloc[0]
    if pct_emp > semaforo_rojo:
        registrar_alerta("CRITICO", "V1-KPI1",
                         f"% Riesgo A+MA supera umbral rojo ({semaforo_rojo}%)",
                         f"{pct_emp:.1f}%", f"≤{semaforo_rojo}%", "Intervención prioritaria. Revisar plan de acción Res. 2764/2022")
    elif pct_emp > 15:
        registrar_alerta("ADVERTENCIA", "V1-KPI1",
                         f"% Riesgo A+MA en zona amarilla (15-35%)",
                         f"{pct_emp:.1f}%", f"<15% verde", "Monitoreo y medidas preventivas")
    else:
        registrar_alerta("INFO", "V1-KPI1",
                         f"% Riesgo A+MA en zona verde (<15%)",
                         f"{pct_emp:.1f}%", f"<15%", "Sin acción urgente")

# -- Alerta: Dimensiones críticas vs Colombia ----------------------------------
if not bench_a.empty and "diferencia_pp" in bench_a.columns:
    dims_criticas = bench_a[
        (bench_a["nivel_analisis"] == "dimension") &
        (bench_a["diferencia_pp"] > 0)
    ]
    n_dims_sobre = len(dims_criticas)
    if n_dims_sobre > 5:
        registrar_alerta("CRITICO", "V1-DIMS",
                         f"{n_dims_sobre} dimensiones por encima del referente nacional",
                         n_dims_sobre, "≤3", "Priorizar intervención en dimensiones críticas señaladas en S7/S8")
    elif n_dims_sobre > 2:
        registrar_alerta("ADVERTENCIA", "V1-DIMS",
                         f"{n_dims_sobre} dimensiones por encima del referente nacional",
                         n_dims_sobre, "≤3", "Revisar plan de intervención para dimensiones críticas")
    else:
        registrar_alerta("INFO", "V1-DIMS",
                         f"Solo {n_dims_sobre} dimensiones sobre el referente nacional",
                         n_dims_sobre, "≤3", "Sin acción urgente")

# -- Alerta: Ausentismo --------------------------------------------------------
if not aus_a.empty:
    dias_tot = aus_a["dias_ausencia"].sum() if "dias_ausencia" in aus_a.columns else 0
    registrar_alerta("INFO", "V3-AUS",
                     f"Ausentismo total registrado para la empresa",
                     f"{dias_tot} días", "—", "Verificar completitud del registro de ausentismo")
    if len(aus_a) < n_trab * 0.1:
        registrar_alerta("ADVERTENCIA", "V3-AUS-COBERT",
                         "Cobertura de ausentismo muy baja (<10% de trabajadores con registro)",
                         f"{len(aus_a)} registros vs {n_trab} trabajadores",
                         "≥10% cobertura", "Solicitar registro completo de ausentismo al cliente")
else:
    registrar_alerta("ADVERTENCIA", "V3-AUS-VACIO",
                     "Sin registros de ausentismo para ACVICOL",
                     "0 registros", "≥1 registro", "Cálculo R10 de costos se basa en estimados; solicitar datos de ausentismo")

# -- Alerta: Protocolos urgentes V2 -------------------------------------------
if not g_prot_a.empty and "categoria_urgencia" in g_prot_a.columns:
    urgentes = g_prot_a[g_prot_a["categoria_urgencia"].str.contains("urgente", case=False, na=False)]
    if len(urgentes) > 0:
        registrar_alerta("CRITICO", "V2-PROT",
                         f"{len(urgentes)} protocolo(s) de INTERVENCIÓN URGENTE identificados",
                         len(urgentes), "0 protocolos urgentes",
                         f"Activar protocolos: {', '.join(urgentes['protocolo_nombre'].astype(str).tolist()[:3])}")

# -- Alerta: Vigilancia epidemiológica -----------------------------------------
if not g_vig_tr_a.empty and "cedula" in g_vig_tr_a.columns:
    n_vig = g_vig_tr_a["cedula"].nunique()
    pct_vig = 100 * n_vig / n_trab if n_trab > 0 else 0
    if pct_vig > 20:
        registrar_alerta("CRITICO", "V2-VIG",
                         f"{n_vig} trabajadores ({pct_vig:.1f}%) cumplen criterios de vigilancia epidemiológica",
                         f"{n_vig} trab ({pct_vig:.1f}%)", "<10%",
                         "Activar programa de vigilancia epidemiológica activa")
    elif pct_vig > 0:
        registrar_alerta("ADVERTENCIA", "V2-VIG",
                         f"{n_vig} trabajadores ({pct_vig:.1f}%) con criterios de vigilancia",
                         f"{n_vig} trab ({pct_vig:.1f}%)", "<10%",
                         "Revisar casos en hoja V2_Vigilancia")

# -- Alerta: Integridad de datos -----------------------------------------------
n_resp_esperadas = n_trab * 422  # aprox ítems completos
if n_respuestas < n_resp_esperadas * 0.70:
    registrar_alerta("ADVERTENCIA", "R1-INTEG",
                     "Volumen de respuestas inferior al 70% del esperado (posibles ítems no respondidos)",
                     n_respuestas, f"≥{int(n_resp_esperadas * 0.70)}", "Revisar tasa de completitud por instrumento")
else:
    registrar_alerta("INFO", "R1-INTEG",
                     "Volumen de respuestas dentro de rango esperado",
                     n_respuestas, f"~{n_resp_esperadas}", "Sin acción")

# ── RESUMEN EJECUTIVO: calcular indicadores síntesis ─────────────────────────
resumen_rows = []

# -- Demografía básica --
if not demog_a.empty:
    n_fem = demog_a[demog_a["sexo"].str.upper().str.strip() == "FEMENINO"].shape[0] if "sexo" in demog_a.columns else 0
    n_mas = demog_a[demog_a["sexo"].str.upper().str.strip() == "MASCULINO"].shape[0] if "sexo" in demog_a.columns else 0
    edad_prom = demog_a["edad_cumplida"].mean() if "edad_cumplida" in demog_a.columns else None
else:
    n_fem, n_mas, edad_prom = 0, 0, None

# -- Riesgo global (baremos factores) --
if not bar_a.empty:
    factores = bar_a[bar_a["nivel_analisis"] == "factor"] if "nivel_analisis" in bar_a.columns else pd.DataFrame()
    if not factores.empty:
        intra_levels = factores[factores["instrumento"].isin(["IntraA", "IntraB"])]
        pct_alto_intra = 0
        if not intra_levels.empty and "nivel_riesgo" in intra_levels.columns:
            n_alto_intra = intra_levels[intra_levels["nivel_riesgo"] >= 4]["cedula"].nunique()
            pct_alto_intra = round(100 * n_alto_intra / n_trab, 1) if n_trab > 0 else 0

# -- Benchmarking sector --
ref_sector_label = f"{ref_sector:.1f}% (referente {sector})"

resumen_rows = [
    ("EMPRESA", "Nombre empresa", EMPRESA),
    ("EMPRESA", "Sector económico (raw)", sector_raw),
    ("EMPRESA", "Sector homologado (ENCST)", sector),
    ("EMPRESA", "N° total trabajadores evaluados", n_trab),
    ("EMPRESA", "Forma A (Jefes)", n_forma_a),
    ("EMPRESA", "Forma B (Colaboradores)", n_forma_b),
    ("EMPRESA", "N° registros de respuesta", n_respuestas),
    ("EMPRESA", "N° registros con ausentismo", len(aus_a)),
    ("DEMOGRAFÍA", "N° Mujeres", n_fem),
    ("DEMOGRAFÍA", "N° Hombres", n_mas),
    ("DEMOGRAFÍA", "Edad promedio", f"{edad_prom:.1f} años" if edad_prom else "N/A"),
    ("V1 — RIESGO PSICOSOCIAL", "Referente sector (% Alto+MuyAlto Intra)", ref_sector_label),
    ("V1 — RIESGO PSICOSOCIAL", "Semáforo rojo si empresa supera", f"{semaforo_rojo}%"),
    ("V1 — RIESGO PSICOSOCIAL", "N° dimensiones sobre referente Colombia",
     str(n_dims_sobre) if 'n_dims_sobre' in dir() else "Ver hoja 06_Benchmarking"),
    ("V1 — RIESGO PSICOSOCIAL", "Baremos aplicados", "Res. 2764/2022 (Ministerio Trabajo Colombia)"),
    ("V1 — RIESGO PSICOSOCIAL", "Pasos aplicados en pipeline", "Pasos 1-21 (Scripts 01, 02a, 02b, 06, 07, 08)"),
    ("V2 — GESTIÓN SALUD MENTAL", "Instrumentos V2", "Afrontamiento, Capital Psicológico, Extralaboral, Estrés"),
    ("V2 — GESTIÓN SALUD MENTAL", "N° indicadores de gestión", str(g_indic_a["indicador"].nunique() if not g_indic_a.empty and "indicador" in g_indic_a.columns else "N/A")),
    ("V2 — GESTIÓN SALUD MENTAL", "N° protocolos priorizados", str(len(g_prot_a)) if not g_prot_a.empty else "N/A"),
    ("V2 — GESTIÓN SALUD MENTAL", "N° trabajadores en vigilancia epidemiológica", str(n_vig) if 'n_vig' in dir() else "Ver hoja V2_Vigilancia"),
    ("V3 — GERENCIAL / ASIS", "KPIs gerenciales calculados", str(len(v3_kpis_a)) if not v3_kpis_a.empty else "N/A"),
    ("V3 — GERENCIAL / ASIS", "SMLV mensual (config)", f"${CFG['smlv_mensual']:,} COP"),
    ("V3 — GERENCIAL / ASIS", "Factor presentismo (config)", str(CFG.get("presentismo_factor", 1.4))),
    ("V3 — GERENCIAL / ASIS", "Costo empleador % (config)", str(CFG.get("costo_empleador_pct", 0.60))),
    ("V3 — GERENCIAL / ASIS", "% imputación psicosocial (config)", str(CFG.get("psicosocial_pct", 0.30))),
    ("PIPELINE", "Total parquets procesados", "35 tablas (dims + facts V1/V2/V3)"),
    ("PIPELINE", "Regla confidencialidad mínimo N", str(CFG.get("confidencialidad_n_min", 5))),
    ("PIPELINE", "Marco normativo", "Resolución 2764/2022 — Ministerio Trabajo Colombia"),
    ("PIPELINE", "Fuentes referencia nacional", "II y III ENCST 2013-2021"),
    ("AUDITORÍA", "Fecha de auditoría", FECHA_AUDITORIA),
    ("AUDITORÍA", "Total alertas generadas", str(len(alertas))),
    ("AUDITORÍA", "Alertas CRITICAS", str(sum(1 for a in alertas if a["Nivel"] == "CRITICO"))),
    ("AUDITORÍA", "Alertas ADVERTENCIA", str(sum(1 for a in alertas if a["Nivel"] == "ADVERTENCIA"))),
    ("AUDITORÍA", "Alertas INFO", str(sum(1 for a in alertas if a["Nivel"] == "INFO"))),
]


# ══════════════════════════════════════════════════════════════════════════════
# CONSTRUIR LIBRO EXCEL
# ══════════════════════════════════════════════════════════════════════════════
print("Construyendo libro Excel de auditoría...")
wb = Workbook()
wb.remove(wb.active)  # quitar hoja en blanco por defecto


# ─────────────────────────────────────────────────────────────────────────────
# 00 — PORTADA
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("00_PORTADA")
ws.column_dimensions["A"].width = 8
ws.column_dimensions["B"].width = 40
ws.column_dimensions["C"].width = 55

# Mega-título
ws.merge_cells("A1:C1")
ws["A1"].value = "LIBRO DE AUDITORÍA ETL — MentalPRO"
ws["A1"].fill = fill(NAVY)
ws["A1"].font = bold_font(WHITE, 18)
ws["A1"].alignment = center()
ws.row_dimensions[1].height = 45

ws.merge_cells("A2:C2")
ws["A2"].value = f"EMPRESA: {EMPRESA}   |   Auditor: Sistema Automático MentalPRO   |   Fecha: {FECHA_AUDITORIA}"
ws["A2"].fill = fill(GOLD)
ws["A2"].font = bold_font(WHITE, 12)
ws["A2"].alignment = center()
ws.row_dimensions[2].height = 25

ws.merge_cells("A3:C3")
ws["A3"].value = "Resolución 2764/2022 — Ministerio de Trabajo Colombia | II y III ENCST 2013-2021"
ws["A3"].fill = fill(CYAN)
ws["A3"].font = bold_font(NAVY, 10)
ws["A3"].alignment = center()
ws.row_dimensions[3].height = 18

# Datos empresa (row 5+)
encabezados = [("CAMPO", "VALOR")]
datos_empresa = [
    ("Empresa auditada", EMPRESA),
    ("Sector económico (raw)", sector_raw),
    ("Sector homologado ENCST", sector),
    ("N° trabajadores evaluados", n_trab),
    ("  — Forma A (Jefes/Directivos)", n_forma_a),
    ("  — Forma B (Colaboradores)", n_forma_b),
    ("N° respuestas totales procesadas", n_respuestas),
    ("Fecha de auditoría", FECHA_AUDITORIA),
    ("Pipeline ETL — Scripts", "01 → 02a → 02b → 06 → 07 → 08 (V1) + 03→07 (V2) + 09 (V3)"),
    ("Framework normativo", "Res. 2764/2022 — Batería de Riesgo Psicosocial Colombia"),
    ("Referente nacional", "III ENCST 2021 (sector) + II-III ENCST 2013-2021 (dominios/dims)"),
    ("Confidencialidad mínimo N", CFG.get("confidencialidad_n_min", 5)),
    ("SMLV mensual (R10)", f"${CFG['smlv_mensual']:,} COP"),
    ("Factor presentismo (R10)", CFG.get("presentismo_factor", 1.4)),
    ("Costo empleador % (R10)", f"{CFG.get('costo_empleador_pct', 0.60)*100:.0f}%"),
    ("% Imputación psicosocial (R10)", f"{CFG.get('psicosocial_pct', 0.30)*100:.0f}%"),
    ("Umbral semáforo ROJO", f">{CFG.get('semaforo_rojo_pct', 35)}%"),
    ("Umbral semáforo AMARILLO", "15-34%"),
    ("Umbral semáforo VERDE", "<15%"),
]

for ri, (campo, valor) in enumerate(datos_empresa, 5):
    bg = GRAY_L if ri % 2 == 0 else WHITE
    ws.cell(row=ri, column=1).value = ""
    c_campo = ws.cell(row=ri, column=2)
    c_valor = ws.cell(row=ri, column=3)
    c_campo.value = campo
    c_valor.value = valor
    for cell in (c_campo, c_valor):
        cell.fill = fill(bg)
        cell.font = std_font(size=10)
        cell.alignment = left_align()
        cell.border = thin_border()
    c_campo.font = bold_font(size=10)
    ws.row_dimensions[ri].height = 16

# Encabezado de la tabla
ws.cell(row=4, column=2).value = "PARÁMETRO"
ws.cell(row=4, column=3).value = "VALOR"
for col in (2, 3):
    ws.cell(row=4, column=col).fill = fill(NAVY)
    ws.cell(row=4, column=col).font = bold_font(WHITE, 10)
    ws.cell(row=4, column=col).alignment = center()
    ws.cell(row=4, column=col).border = thin_border()
ws.row_dimensions[4].height = 20

# Índice de hojas (fila 26+)
row_idx = len(datos_empresa) + 7
ws.merge_cells(f"A{row_idx}:C{row_idx}")
ws[f"A{row_idx}"].value = "ÍNDICE DE HOJAS DEL LIBRO DE AUDITORÍA"
ws[f"A{row_idx}"].fill = fill(NAVY)
ws[f"A{row_idx}"].font = bold_font(WHITE, 11)
ws[f"A{row_idx}"].alignment = center()
ws.row_dimensions[row_idx].height = 22
row_idx += 1

hojas_indice = [
    ("00_PORTADA",              "Esta hoja — Metadatos y parámetros de auditoría"),
    ("01_ETL_StarSchema",       "Paso 1: Respuestas limpias ACVICOL (R1/R6/R17)"),
    ("02a_Paso1_Codificacion",  "Paso 2: Codificación texto → número (6 escalas)"),
    ("02a_Paso2_Inversion",     "Paso 3: Inversión ítems Nivel 1 (R3, 73+68+23 ítems)"),
    ("02a_Pasos3a8_Agrup",      "Pasos 4-8: Agrupaciones instrumento/dim/dominio/factor"),
    ("02b_Baremos_Dim",         "Pasos 9-11: Baremos dimensiones (riesgo/protección) Res. 2764"),
    ("02b_Baremos_Dom",         "Pasos 12-13: Baremos dominios IntraA/B + Individual"),
    ("02b_Baremos_Factor",      "Pasos 14-15: Baremos factores Intra/Extra/Estrés/Individual"),
    ("06_Benchmarking",         "Pasos 16-19: Comparativa empresa vs sector ENCST y Colombia"),
    ("07_Frecuencias",          "Paso 20: Distribución % respuestas por pregunta"),
    ("07_Top20_Comparables",    "Paso 20b: Top 20 preguntas con mayor diferencia vs ENCST"),
    ("08_Consolidado",          "Paso 21: Fact consolidado con demografía (formato largo)"),
    ("V2_Estandarizado",        "V2 Sub-paso 1: Estandarización 0-1 por ítem"),
    ("V2_Invertido",            "V2 Sub-paso 2: Inversión V2 específica"),
    ("V2_Indicadores",          "V2 Paso 3: Scores por indicador de gestión"),
    ("V2_Lineas",               "V2 Paso 4: Scores por línea de gestión (20 líneas)"),
    ("V2_Ejes",                 "V2 Paso 5: Scores por eje de gestión"),
    ("V2_Niveles_Indicadores",  "V2 Categorización indicadores (5 niveles gestión)"),
    ("V2_Niveles_Lineas",       "V2 Categorización líneas de gestión"),
    ("V2_Niveles_Ejes",         "V2 Categorización ejes de gestión"),
    ("V2_Protocolos",           "V2 Paso 6: Protocolos priorizados por sector/urgencia"),
    ("V2_Vigilancia",           "V2 Paso 6b: Trabajadores con criterios vigilancia epidemiológica"),
    ("V3_KPIs_Globales",        "V3 Paso 1: 19 KPIs ejecutivos gerenciales"),
    ("V3_Demografia",           "V3 Paso 2: Distribución demográfica empresa"),
    ("V3_Costos",               "V3 Paso 3: Cálculo económico ausentismo (R10)"),
    ("V3_Benchmarking",         "V3 Paso 4: Benchmarking ejecutivo sectorial"),
    ("V3_Ranking_Areas",        "V3 Paso 5: Ranking top 5 áreas críticas"),
    ("ZZ_RESUMEN",              "RESUMEN EJECUTIVO consolidado de todos los visualizadores"),
    ("ZZ_ALERTAS",              "ALERTAS y puntos de chequeo para auditoría"),
]

ws.cell(row=row_idx, column=2).value = "Hoja"
ws.cell(row=row_idx, column=3).value = "Contenido"
for col in (2, 3):
    ws.cell(row=row_idx, column=col).fill = fill(GOLD)
    ws.cell(row=row_idx, column=col).font = bold_font(NAVY, 10)
    ws.cell(row=row_idx, column=col).alignment = center()
    ws.cell(row=row_idx, column=col).border = thin_border()
ws.row_dimensions[row_idx].height = 18
row_idx += 1

for i, (hoja, desc) in enumerate(hojas_indice, 0):
    bg = GRAY_L if i % 2 == 0 else WHITE
    ws.cell(row=row_idx + i, column=2).value = hoja
    ws.cell(row=row_idx + i, column=3).value = desc
    for col in (2, 3):
        ws.cell(row=row_idx + i, column=col).fill = fill(bg)
        ws.cell(row=row_idx + i, column=col).font = std_font(size=9)
        ws.cell(row=row_idx + i, column=col).alignment = left_align()
        ws.cell(row=row_idx + i, column=col).border = thin_border()
    ws.row_dimensions[row_idx + i].height = 14

print("  [OK] 00_PORTADA")


# ─────────────────────────────────────────────────────────────────────────────
# 01 — ETL Star Schema (Respuestas limpias)
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("01_ETL_StarSchema")
write_title_block(ws,
    "PASO 1 — ETL Star Schema: Respuestas Limpias ACVICOL",
    "Script: 01_etl_star_schema.py | Reglas: R1, R6, R15, R17")

# Métricas de calidad
row_info = 4
ws.merge_cells(f"A{row_info}:Z{row_info}")
ws[f"A{row_info}"].value = (
    f"Registros cargados: {len(resp_a)}  |  "
    f"Trabajadores únicos: {n_trab}  |  "
    f"Forma A: {n_forma_a}  |  Forma B: {n_forma_b}  |  "
    f"Sector raw: '{sector_raw}'  →  Sector ENCST: '{sector}'  |  "
    f"Registros ausentismo: {len(aus_a)}"
)
ws[f"A{row_info}"].fill = fill(GRAY_M)
ws[f"A{row_info}"].font = bold_font(NAVY, 9)
ws[f"A{row_info}"].alignment = left_align()
ws.row_dimensions[row_info].height = 18

# Resumen por instrumento
row_info2 = 5
ws.merge_cells(f"A{row_info2}:Z{row_info2}")
if not resp_a.empty and "id_pregunta" in resp_a.columns:
    resp_a["instrumento_sufijo"] = resp_a["id_pregunta"].str.split("_").str[-1]
    resumen_inst = resp_a.groupby("instrumento_sufijo").agg(
        n_respuestas=("id_respuesta", "count"),
        n_trabajadores=("cedula", "nunique"),
        n_items_unicos=("id_pregunta", "nunique")
    ).reset_index()
    inst_str = " | ".join([
        f"{r['instrumento_sufijo'].upper()}: {r['n_respuestas']} resp / {r['n_trabajadores']} trab / {r['n_items_unicos']} ítems"
        for _, r in resumen_inst.iterrows()
    ])
else:
    inst_str = "No disponible"

ws[f"A{row_info2}"].value = f"Desglose por instrumento → {inst_str}"
ws[f"A{row_info2}"].fill = fill(GRAY_L)
ws[f"A{row_info2}"].font = std_font(NAVY, 8)
ws[f"A{row_info2}"].alignment = left_align()
ws.row_dimensions[row_info2].height = 16

# Datos
cols_resp = ["empresa", "cedula", "forma_intra", "sector_economico", "sector_rag",
             "id_pregunta", "id_respuesta"]
cols_resp = [c for c in cols_resp if c in resp_a.columns]
write_df_to_ws(ws, resp_a[cols_resp].head(5000), start_row=6)
print("  [OK] 01_ETL_StarSchema")


# ─────────────────────────────────────────────────────────────────────────────
# 02a — PASO 1: Codificación texto → número
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("02a_Paso1_Codificacion")
write_title_block(ws,
    "PASO 2 — Codificación: Texto → Valor Numérico",
    "Script: 02a_scoring_bateria.py | Reglas: R1, R15 | 6 escalas de codificación")

# Tabla de escalas de codificación (referencia)
escalas = [
    ("ESCALA", "INSTRUMENTO", "RESPUESTA TEXTO", "VALOR NUMÉRICO"),
    ("Likert 0-4", "IntraA/B, Extralaboral", "Siempre", "4"),
    ("Likert 0-4", "IntraA/B, Extralaboral", "Casi siempre", "3"),
    ("Likert 0-4", "IntraA/B, Extralaboral", "Algunas veces", "2"),
    ("Likert 0-4", "IntraA/B, Extralaboral", "Casi nunca", "1"),
    ("Likert 0-4", "IntraA/B, Extralaboral", "Nunca", "0"),
    ("Dicotómica", "IntraA ítems 106/116; IntraB ítem 89", "Sí", "1"),
    ("Dicotómica", "IntraA ítems 106/116; IntraB ítem 89", "No", "0"),
    ("Estrés G1 (9/6/3/0)", "Estrés ítems 1,2,3,9,13,14,15,23,24", "Siempre / Casi siempre / A veces / Nunca", "9 / 6 / 3 / 0"),
    ("Estrés G2 (6/4/2/0)", "Estrés ítems 4,5,6,10,11,16,17,18,19,25,26,27,28", "Siempre / Casi siempre / A veces / Nunca", "6 / 4 / 2 / 0"),
    ("Estrés G3 (3/2/1/0)", "Estrés ítems 7,8,12,20,21,22,29,30,31", "Siempre / Casi siempre / A veces / Nunca", "3 / 2 / 1 / 0"),
    ("Afrontamiento (0/0.5/0.7/1)", "Afrontamiento (ítems 1-12)", "Nunca / A veces / Frecuentemente / Siempre", "0 / 0.5 / 0.7 / 1.0"),
    ("Capital Psicológico (0/0.5/0.7/1)", "CapPsico (ítems 1-12)", "Tot. desacuerdo / Desacuerdo / De acuerdo / Tot. acuerdo", "0 / 0.5 / 0.7 / 1.0"),
]

for ri, row_data in enumerate(escalas, 4):
    bg = NAVY if ri == 4 else (GRAY_L if ri % 2 == 0 else WHITE)
    for ci, val in enumerate(row_data, 1):
        cell = ws.cell(row=ri, column=ci)
        cell.value = val
        cell.fill = fill(bg)
        cell.font = bold_font(WHITE, 9) if ri == 4 else std_font(size=9)
        cell.alignment = left_align()
        cell.border = thin_border()
    ws.row_dimensions[ri].height = 16

ws.merge_cells(f"A{len(escalas)+6}:Z{len(escalas)+6}")
ws[f"A{len(escalas)+6}"].value = "DATOS: fact_scores_brutos — columna valor_numerico (filtrado ACVICOL, primeras 5000 filas)"
ws[f"A{len(escalas)+6}"].fill = fill(GOLD)
ws[f"A{len(escalas)+6}"].font = bold_font(NAVY, 9)
ws[f"A{len(escalas)+6}"].alignment = left_align()
ws.row_dimensions[len(escalas)+6].height = 16

cols_bruto = ["cedula", "empresa", "forma_intra", "id_pregunta", "id_respuesta", "valor_numerico"]
cols_bruto = [c for c in cols_bruto if c in bruto_a.columns]
write_df_to_ws(ws, bruto_a[cols_bruto].head(5000), start_row=len(escalas)+7)
print("  [OK] 02a_Paso1_Codificacion")


# ─────────────────────────────────────────────────────────────────────────────
# 02a — PASO 2: Inversión de ítems (R3)
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("02a_Paso2_Inversion")
write_title_block(ws,
    "PASO 3 — Inversión de Ítems Nivel 1 (R3)",
    "Script: 02a_scoring_bateria.py | Fórmula Likert: valor_invertido = 4 - valor_numerico")

# Resumen inversión
inv_info = [
    ("INSTRUMENTO", "N° ÍTEMS INVERTIDOS", "FÓRMULA", "NOTA"),
    ("Intralaboral A", "73 ítems", "valor_invertido = 4 − valor_numerico", "Ítems: 4,5,6,9,12,14,32,34,39-51,53-105..."),
    ("Intralaboral B", "68 ítems", "valor_invertido = 4 − valor_numerico", "Ítems: 4,5,6,9,12,14,22,24,29-65,67-88,98"),
    ("Extralaboral", "23 ítems", "valor_invertido = 4 − valor_numerico", "Ítems: 1,4,5,7-23,25,27,29"),
    ("Afrontamiento ítems 5-8", "4 ítems", "Escala propia invertida", "Nunca→1.0, A veces→0.7, Frecuentemente→0.5, Siempre→0.0"),
    ("Estrés", "0 ítems", "Sin inversión Nivel 1", "valor_invertido = valor_numerico"),
    ("Capital Psicológico", "0 ítems", "Sin inversión Nivel 1", "valor_invertido = valor_numerico"),
]
for ri, row_data in enumerate(inv_info, 4):
    bg = NAVY if ri == 4 else (GRAY_L if ri % 2 == 0 else WHITE)
    for ci, val in enumerate(row_data, 1):
        cell = ws.cell(row=ri, column=ci)
        cell.value = val
        cell.fill = fill(bg)
        cell.font = bold_font(WHITE, 9) if ri == 4 else std_font(size=9)
        cell.alignment = left_align()
        cell.border = thin_border()
    ws.row_dimensions[ri].height = 16

# Métricas de inversión
n_invertidos = 0
if not bruto_a.empty and "valor_numerico" in bruto_a.columns and "valor_invertido" in bruto_a.columns:
    diff = bruto_a["valor_numerico"] != bruto_a["valor_invertido"]
    n_invertidos = diff.sum()

ws.merge_cells(f"A{len(inv_info)+6}:Z{len(inv_info)+6}")
ws[f"A{len(inv_info)+6}"].value = (
    f"RESULTADO: {n_invertidos:,} ítems con inversión aplicada sobre {len(bruto_a):,} registros ACVICOL "
    f"({100*n_invertidos/len(bruto_a):.1f}% invertidos)" if len(bruto_a) > 0 else "Sin datos"
)
ws[f"A{len(inv_info)+6}"].fill = fill(CYAN)
ws[f"A{len(inv_info)+6}"].font = bold_font(NAVY, 9)
ws.row_dimensions[len(inv_info)+6].height = 16

cols_inv = ["cedula", "forma_intra", "id_pregunta", "instrumento",
            "valor_numerico", "valor_invertido", "max_item_score"]
cols_inv = [c for c in cols_inv if c in bruto_a.columns]
write_df_to_ws(ws, bruto_a[cols_inv].head(5000), start_row=len(inv_info)+7)
print("  [OK] 02a_Paso2_Inversion")


# ─────────────────────────────────────────────────────────────────────────────
# 02a — PASOS 3-8: Agrupaciones
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("02a_Pasos3a8_Agrup")
write_title_block(ws,
    "PASOS 4-8 — Agrupaciones: instrumento / dimensión / dominio / factor",
    "Script: 02a_scoring_bateria.py | JOIN con dim_pregunta construida internamente")

# Info
ws.merge_cells("A4:Z4")
ws["A4"].value = (
    "Cada ítem se asigna a: instrumento (IntraA/B, Extralaboral, Estres, Afrontamiento, CapPsico) "
    "→ dimensión (19 IntraA, 16 IntraB, 7 Extra, etc.) → dominio → factor. "
    "dim_pregunta se construye internamente desde categorias_analisis con SWAP de sufijos afrontamiento/capitalpsicologico."
)
ws["A4"].fill = fill(GRAY_M)
ws["A4"].font = std_font(NAVY, 9)
ws["A4"].alignment = left_align()
ws.row_dimensions[4].height = 28

cols_agr = ["cedula", "forma_intra", "id_pregunta", "instrumento",
            "dimension", "dominio", "factor", "valor_invertido"]
cols_agr = [c for c in cols_agr if c in bruto_a.columns]
write_df_to_ws(ws, bruto_a[cols_agr].head(5000), start_row=5)
print("  [OK] 02a_Pasos3a8_Agrup")


# ─────────────────────────────────────────────────────────────────────────────
# 02b — BAREMOS DIMENSIONES (Pasos 9-11)
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("02b_Baremos_Dim")
write_title_block(ws,
    "PASOS 9-11 — Baremos: Dimensiones (Riesgo + Protección)",
    "Script: 02b_baremos.py | Res. 2764/2022 + Avantum | Fórmula: PT = round(bruto/max × 100, 1)")

ws.merge_cells("A4:Z4")
ws["A4"].value = (
    "Fórmula transformación: puntaje_transformado = round(suma_bruto / max_posible × 100, 1)  |  "
    "5 niveles: Sin riesgo / Riesgo bajo / Riesgo medio / Riesgo alto / Riesgo muy alto  |  "
    "Tipos baremo: 'riesgo' (intra/extra), 'afrontamiento_dim', 'capitalpsicologico_dim'  |  "
    "nivel_analisis='dimension'"
)
ws["A4"].fill = fill(GRAY_M)
ws["A4"].font = std_font(NAVY, 8)
ws["A4"].alignment = left_align()
ws.row_dimensions[4].height = 20

if not bar_a.empty:
    df_dim = bar_a[bar_a["nivel_analisis"] == "dimension"].copy() if "nivel_analisis" in bar_a.columns else bar_a.copy()
    cols_dim = ["cedula", "empresa", "forma_intra", "instrumento",
                "nivel_analisis", "nombre_nivel", "puntaje_bruto",
                "transformacion_max", "puntaje_transformado",
                "nivel_riesgo", "etiqueta_nivel", "tipo_baremo"]
    cols_dim = [c for c in cols_dim if c in df_dim.columns]
    write_df_to_ws(ws, df_dim[cols_dim], start_row=5, col_color_fn=color_by_nivel_riesgo)
else:
    ws.cell(row=5, column=1).value = "Sin datos de baremos para ACVICOL"
print("  [OK] 02b_Baremos_Dim")


# ─────────────────────────────────────────────────────────────────────────────
# 02b — BAREMOS DOMINIOS (Pasos 12-13)
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("02b_Baremos_Dom")
write_title_block(ws,
    "PASOS 12-13 — Baremos: Dominios",
    "Script: 02b_baremos.py | Dominios IntraA/B (4 c/u) + Dominio CapPsico")

if not bar_a.empty:
    df_dom = bar_a[bar_a["nivel_analisis"] == "dominio"].copy() if "nivel_analisis" in bar_a.columns else pd.DataFrame()
    cols_dom = ["cedula", "empresa", "forma_intra", "instrumento",
                "nivel_analisis", "nombre_nivel", "puntaje_bruto",
                "transformacion_max", "puntaje_transformado",
                "nivel_riesgo", "etiqueta_nivel", "tipo_baremo"]
    cols_dom = [c for c in cols_dom if c in df_dom.columns]
    write_df_to_ws(ws, df_dom[cols_dom], start_row=4, col_color_fn=color_by_nivel_riesgo)
else:
    ws.cell(row=4, column=1).value = "Sin datos"
print("  [OK] 02b_Baremos_Dom")


# ─────────────────────────────────────────────────────────────────────────────
# 02b — BAREMOS FACTORES (Pasos 14-15)
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("02b_Baremos_Factor")
write_title_block(ws,
    "PASOS 14-15 — Baremos: Factores (Intra / Extra / Estrés / Individual)",
    "Script: 02b_baremos.py | Paso 14.1: Estrés con 4 promedios ponderados (max=61.16)")

ws.merge_cells("A4:Z4")
ws["A4"].value = (
    "Factores: IntraA (max variable por forma), IntraB, Extralaboral (ambas formas), "
    "Estrés (4 grupos ponderados, max=61.16), Individual (Afrontamiento+CapPsico, max=24)  |  "
    "nivel_analisis='factor'"
)
ws["A4"].fill = fill(GRAY_M)
ws["A4"].font = std_font(NAVY, 8)
ws["A4"].alignment = left_align()
ws.row_dimensions[4].height = 20

if not bar_a.empty:
    df_fac = bar_a[bar_a["nivel_analisis"] == "factor"].copy() if "nivel_analisis" in bar_a.columns else pd.DataFrame()
    cols_fac = ["cedula", "empresa", "forma_intra", "instrumento",
                "nivel_analisis", "nombre_nivel", "puntaje_bruto",
                "transformacion_max", "puntaje_transformado",
                "nivel_riesgo", "etiqueta_nivel", "tipo_baremo"]
    cols_fac = [c for c in cols_fac if c in df_fac.columns]
    write_df_to_ws(ws, df_fac[cols_fac], start_row=5, col_color_fn=color_by_nivel_riesgo)
else:
    ws.cell(row=5, column=1).value = "Sin datos"
print("  [OK] 02b_Baremos_Factor")


# ─────────────────────────────────────────────────────────────────────────────
# 06 — BENCHMARKING (Pasos 16-19)
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("06_Benchmarking")
write_title_block(ws,
    "PASOS 16-19 — Benchmarking: Empresa vs Sector y Colombia",
    "Script: 06_benchmarking.py | ENCST III 2021 (sector) + II-III 2013-2021 (dominios/dims)")

ws.merge_cells("A4:Z4")
ws["A4"].value = (
    f"Referente sector ACVICOL: '{sector}' = {ref_sector:.1f}% (III ENCST 2021)  |  "
    "Paso 16: Riesgo total empresa + flag reevaluación  |  "
    "Paso 17: % A+MA IntraA/B vs sector  |  Paso 18: % A+MA dominios vs Colombia  |  "
    "Paso 19: % A+MA dimensiones (11 comparables) vs ENCST  |  "
    "Semáforo: rojo si diferencia_pp > 0 (empresa peor que referente)"
)
ws["A4"].fill = fill(GRAY_M)
ws["A4"].font = std_font(NAVY, 8)
ws["A4"].alignment = left_align()
ws.row_dimensions[4].height = 28

cols_bench = ["empresa", "sector_rag", "nivel_analisis", "nombre_nivel",
              "forma_intra", "instrumento", "n_total", "n_alto",
              "pct_empresa", "pct_referencia", "tipo_referencia",
              "diferencia_pp", "semaforo"]
cols_bench = [c for c in cols_bench if c in bench_a.columns]
write_df_to_ws(ws, bench_a[cols_bench], start_row=5, col_color_fn=color_by_nivel_riesgo)
print("  [OK] 06_Benchmarking")


# ─────────────────────────────────────────────────────────────────────────────
# 07 — FRECUENCIAS (Paso 20)
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("07_Frecuencias")
write_title_block(ws,
    "PASO 20 — Frecuencias de Respuestas por Pregunta",
    "Script: 07_frecuencias_preguntas.py | R8: n<5 → pct=None (Confidencial)")

if not freq_a.empty:
    cols_frq = [c for c in freq_a.columns if c in [
        "empresa", "forma_intra", "id_pregunta", "id_respuesta",
        "n_respuesta", "n_total", "pct_respuesta", "instrumento", "dimension"
    ]]
    write_df_to_ws(ws, freq_a[cols_frq].head(3000), start_row=4)
else:
    ws.cell(row=4, column=1).value = "Sin datos"
print("  [OK] 07_Frecuencias")


# ─────────────────────────────────────────────────────────────────────────────
# 07 — TOP 20 COMPARABLES (Paso 20b)
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("07_Top20_Comparables")
write_title_block(ws,
    "PASO 20b — Top 20 Preguntas con Mayor Diferencia vs ENCST",
    "Script: 07_frecuencias_preguntas.py | 39 preguntas comparables con referente nacional")

if not top20_a.empty:
    cols_top = [c for c in top20_a.columns]
    write_df_to_ws(ws, top20_a[cols_top], start_row=4, col_color_fn=color_by_nivel_riesgo)
else:
    ws.cell(row=4, column=1).value = "Sin datos"
print("  [OK] 07_Top20_Comparables")


# ─────────────────────────────────────────────────────────────────────────────
# 08 — CONSOLIDADO (Paso 21)
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("08_Consolidado")
write_title_block(ws,
    "PASO 21 — Consolidado: Baremo + Demografía (Formato Largo)",
    "Script: 08_consolidacion.py | fact_scores_baremo × dim_trabajador × dim_demografia × dim_ausentismo")

ws.merge_cells("A4:Z4")
ws["A4"].value = (
    "Enriquecimiento: 1 fila por trabajador × nivel_analisis × nombre_nivel. "
    "Agrega: área, cargo, sexo, edad, antigüedad, días_ausencia (LEFT JOIN ausentismo — R6). "
    f"Total registros ACVICOL: {len(cons_a):,}"
)
ws["A4"].fill = fill(GRAY_M)
ws["A4"].font = std_font(NAVY, 8)
ws["A4"].alignment = left_align()
ws.row_dimensions[4].height = 20

if not cons_a.empty:
    cols_cons = ["cedula", "empresa", "forma_intra", "area_departamento",
                 "categoria_cargo", "instrumento", "nivel_analisis", "nombre_nivel",
                 "puntaje_transformado", "nivel_riesgo", "etiqueta_nivel",
                 "sexo", "edad_cumplida", "antiguedad_empresa_años_cumplidos",
                 "dias_ausencia"]
    cols_cons = [c for c in cols_cons if c in cons_a.columns]
    write_df_to_ws(ws, cons_a[cols_cons].head(5000), start_row=5, col_color_fn=color_by_nivel_riesgo)
else:
    ws.cell(row=5, column=1).value = "Sin datos"
print("  [OK] 08_Consolidado")


# ─────────────────────────────────────────────────────────────────────────────
# V2 — ESTANDARIZADO (Sub-paso 1)
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("V2_Estandarizado")
write_title_block(ws,
    "V2 Sub-paso 1 — Estandarización 0-1",
    "Script: 03_scoring_gestion.py | 4→1.00, 3→0.75, 2→0.50, 1→0.25, 0→0.00")

if not g_estand_a.empty:
    cols_est = ["cedula", "forma_intra", "id_pregunta", "valor_numerico",
                "valor_estandarizado"] if "valor_estandarizado" in g_estand_a.columns else list(g_estand_a.columns)
    cols_est = [c for c in cols_est if c in g_estand_a.columns]
    write_df_to_ws(ws, g_estand_a[cols_est].head(3000), start_row=4)
else:
    ws.cell(row=4, column=1).value = "Sin datos"
print("  [OK] V2_Estandarizado")


# ─────────────────────────────────────────────────────────────────────────────
# V2 — INVERTIDO (Sub-paso 2)
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("V2_Invertido")
write_title_block(ws,
    "V2 Sub-paso 2 — Inversión V2 Específica",
    "Script: 03_scoring_gestion.py | Ítems autonegación, evitación cognitiva/conductual, etc.")

if not g_invert_a.empty:
    cols_inv2 = list(g_invert_a.columns)[:10]
    write_df_to_ws(ws, g_invert_a[cols_inv2].head(3000), start_row=4)
else:
    ws.cell(row=4, column=1).value = "Sin datos"
print("  [OK] V2_Invertido")


# ─────────────────────────────────────────────────────────────────────────────
# V2 — INDICADORES
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("V2_Indicadores")
write_title_block(ws,
    "V2 Paso 3 — Scores por Indicador de Gestión",
    "Script: 03_scoring_gestion.py + 04_categorias_gestion.py | 61 indicadores (ítem→indicador→línea→eje)")

if not g_indic_a.empty:
    write_df_to_ws(ws, g_indic_a.head(3000), start_row=4, col_color_fn=color_gestion)
else:
    ws.cell(row=4, column=1).value = "Sin datos"
print("  [OK] V2_Indicadores")


# ─────────────────────────────────────────────────────────────────────────────
# V2 — LÍNEAS
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("V2_Lineas")
write_title_block(ws,
    "V2 Paso 4 — Scores por Línea de Gestión (20 líneas)",
    "Script: 04_categorias_gestion.py | Umbrales: >0.79=Prorrogable, 0.65-0.79=Preventiva, 0.45-0.65=Mejora, 0.29-0.45=Correctiva, ≤0.29=Urgente")

if not g_lineas_a.empty:
    write_df_to_ws(ws, g_lineas_a.head(2000), start_row=4, col_color_fn=color_gestion)
else:
    ws.cell(row=4, column=1).value = "Sin datos"
print("  [OK] V2_Lineas")


# ─────────────────────────────────────────────────────────────────────────────
# V2 — EJES
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("V2_Ejes")
write_title_block(ws,
    "V2 Paso 5 — Scores por Eje de Gestión",
    "Script: 04_categorias_gestion.py | Ejes: Afrontamiento, Emocional, Cognitivo, Extralaboral")

if not g_ejes_a.empty:
    write_df_to_ws(ws, g_ejes_a.head(1000), start_row=4, col_color_fn=color_gestion)
else:
    ws.cell(row=4, column=1).value = "Sin datos"
print("  [OK] V2_Ejes")


# ─────────────────────────────────────────────────────────────────────────────
# V2 — NIVELES INDICADORES
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("V2_Niveles_Indicadores")
write_title_block(ws,
    "V2 — Categorización: Niveles por Indicador",
    "Script: 04_categorias_gestion.py | nivel_gestion: Urgente / Correctiva / Mejora / Preventiva / Prorrogable")

if not g_niv_ind_a.empty:
    write_df_to_ws(ws, g_niv_ind_a, start_row=4, col_color_fn=color_gestion)
else:
    ws.cell(row=4, column=1).value = "Sin datos"
print("  [OK] V2_Niveles_Indicadores")


# ─────────────────────────────────────────────────────────────────────────────
# V2 — NIVELES LÍNEAS
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("V2_Niveles_Lineas")
write_title_block(ws,
    "V2 — Categorización: Niveles por Línea de Gestión",
    "Script: 04_categorias_gestion.py")

if not g_niv_lin_a.empty:
    write_df_to_ws(ws, g_niv_lin_a, start_row=4, col_color_fn=color_gestion)
else:
    ws.cell(row=4, column=1).value = "Sin datos"
print("  [OK] V2_Niveles_Lineas")


# ─────────────────────────────────────────────────────────────────────────────
# V2 — NIVELES EJES
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("V2_Niveles_Ejes")
write_title_block(ws,
    "V2 — Categorización: Niveles por Eje de Gestión",
    "Script: 04_categorias_gestion.py")

if not g_niv_eje_a.empty:
    write_df_to_ws(ws, g_niv_eje_a, start_row=4, col_color_fn=color_gestion)
else:
    ws.cell(row=4, column=1).value = "Sin datos"
print("  [OK] V2_Niveles_Ejes")


# ─────────────────────────────────────────────────────────────────────────────
# V2 — PROTOCOLOS
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("V2_Protocolos")
write_title_block(ws,
    "V2 Paso 6 — Protocolos Priorizados por Sector/Urgencia",
    "Script: 05_prioridades_protocolos.py | 20 protocolos (PROT-01 a PROT-20) | Urgencia 70% + Sector 30%")

if not g_prot_a.empty:
    write_df_to_ws(ws, g_prot_a, start_row=4, col_color_fn=color_gestion)
else:
    ws.cell(row=4, column=1).value = "Sin datos de protocolos"
print("  [OK] V2_Protocolos")


# ─────────────────────────────────────────────────────────────────────────────
# V2 — VIGILANCIA EPIDEMIOLÓGICA
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("V2_Vigilancia")
write_title_block(ws,
    "V2 Paso 6b — Vigilancia Epidemiológica",
    "Script: 06_vigilancia_epidemiologica.py | 11 indicadores | Listado nominativo trabajadores")

# Resumen vigilancia
ws.merge_cells("A4:Z4")
n_vig_show = len(g_vig_tr_a) if not g_vig_tr_a.empty else 0
ws["A4"].value = f"Trabajadores con criterios de vigilancia: {n_vig_show} | R8: grupos <5 → Confidencial"
ws["A4"].fill = fill(GRAY_M)
ws["A4"].font = bold_font(NAVY, 9)
ws["A4"].alignment = left_align()
ws.row_dimensions[4].height = 16

# Resumen por indicador
if not g_vig_res_a.empty:
    write_df_to_ws(ws, g_vig_res_a, start_row=5)
elif not g_vig_tr_a.empty:
    write_df_to_ws(ws, g_vig_tr_a.head(500), start_row=5)
else:
    ws.cell(row=5, column=1).value = "Sin datos de vigilancia epidemiológica"
print("  [OK] V2_Vigilancia")


# ─────────────────────────────────────────────────────────────────────────────
# V3 — KPIs GLOBALES
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("V3_KPIs_Globales")
write_title_block(ws,
    "V3 Paso 1 — KPIs Ejecutivos Gerenciales (19 KPIs)",
    "Script: 09_asis_gerencial.py | IAEE, IBET, IBFT, y otros 16 indicadores ejecutivos")

if not v3_kpis_a.empty:
    write_df_to_ws(ws, v3_kpis_a, start_row=4, col_color_fn=color_by_nivel_riesgo)
else:
    ws.cell(row=4, column=1).value = "Sin datos de KPIs V3"
print("  [OK] V3_KPIs_Globales")


# ─────────────────────────────────────────────────────────────────────────────
# V3 — DEMOGRAFÍA
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("V3_Demografia")
write_title_block(ws,
    "V3 Paso 2 — Distribución Demográfica de la Empresa",
    "Script: 09_asis_gerencial.py | Sexo, edad, antigüedad, escolaridad, estado civil, dependientes")

if not v3_demog_a.empty:
    write_df_to_ws(ws, v3_demog_a, start_row=4)
else:
    # Fallback: usar dim_demografia directo
    if not demog_a.empty:
        write_df_to_ws(ws, demog_a, start_row=4)
    else:
        ws.cell(row=4, column=1).value = "Sin datos demográficos"
print("  [OK] V3_Demografia")


# ─────────────────────────────────────────────────────────────────────────────
# V3 — COSTOS (R10)
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("V3_Costos")
write_title_block(ws,
    "V3 Paso 3 — Cálculo Económico de Ausentismo (R10)",
    "Script: 09_asis_gerencial.py | 6 pasos: SMLV → pérdida ausentismo → empleador → presentismo → psicosocial")

# Tabla de fórmula R10
formula_r10 = [
    ("PASO R10", "FÓRMULA", "DESCRIPCIÓN"),
    ("1", f"SMLV_anual = {CFG['smlv_mensual']:,} × 12", f"${CFG['smlv_mensual']*12:,} COP por trabajador/año"),
    ("2", f"capacidad_anual = n_trabajadores × SMLV_anual", f"Capacidad productiva total empresa = {n_trab} × ${CFG['smlv_mensual']*12:,}"),
    ("3", "perdida_ausentismo = (pct_ausentismo / 100) × capacidad_anual", "Pérdida directa por días no trabajados"),
    ("4", f"perdida_empleador = perdida_ausentismo × {CFG.get('costo_empleador_pct', 0.60)}", f"Solo costo a cargo del empleador ({CFG.get('costo_empleador_pct', 0.60)*100:.0f}%)"),
    ("5", f"perdida_productividad = perdida_empleador × {CFG.get('presentismo_factor', 1.40)}", f"Factor presentismo ({CFG.get('presentismo_factor', 1.40)}×) — baja productividad trabajador presente"),
    ("6", f"perdida_psicosocial = perdida_productividad × {CFG.get('psicosocial_pct', 0.30)}", f"Imputación factores psicosociales ({CFG.get('psicosocial_pct', 0.30)*100:.0f}% del costo total)"),
]
for ri, row_data in enumerate(formula_r10, 4):
    bg = NAVY if ri == 4 else (GOLD if ri == 4 else (GRAY_L if ri % 2 == 0 else WHITE))
    if ri == 4:
        bg = NAVY
    for ci, val in enumerate(row_data, 1):
        cell = ws.cell(row=ri, column=ci)
        cell.value = val
        cell.fill = fill(bg)
        cell.font = bold_font(WHITE, 9) if ri == 4 else std_font(size=9)
        cell.alignment = left_align()
        cell.border = thin_border()
    ws.row_dimensions[ri].height = 18

ws.merge_cells(f"A{len(formula_r10)+5}:Z{len(formula_r10)+5}")
ws[f"A{len(formula_r10)+5}"].value = "DATOS REALES CALCULADOS:"
ws[f"A{len(formula_r10)+5}"].fill = fill(GOLD)
ws[f"A{len(formula_r10)+5}"].font = bold_font(NAVY, 10)
ws[f"A{len(formula_r10)+5}"].alignment = left_align()
ws.row_dimensions[len(formula_r10)+5].height = 18

if not v3_cost_a.empty:
    write_df_to_ws(ws, v3_cost_a, start_row=len(formula_r10)+6)
else:
    ws.cell(row=len(formula_r10)+6, column=1).value = (
        f"Nota: Sin registros de ausentismo para ACVICOL. "
        f"Capacidad productiva estimada: ${n_trab * CFG['smlv_mensual'] * 12:,} COP/año"
    )
print("  [OK] V3_Costos")


# ─────────────────────────────────────────────────────────────────────────────
# V3 — BENCHMARKING EJECUTIVO
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("V3_Benchmarking")
write_title_block(ws,
    "V3 Paso 4 — Benchmarking Ejecutivo Sectorial",
    "Script: 09_asis_gerencial.py | Empresa vs sector económico ENCST 2021")

# Tabla referentes por sector
ws.merge_cells("A4:Z4")
ws["A4"].value = (
    "Referentes sectoriales ENCST III 2021: "
    + " | ".join([f"{s}: {v:.1f}%" for s, v in CFG.get("benchmark_sector", {}).items() if not s.startswith("_")])
)
ws["A4"].fill = fill(GRAY_M)
ws["A4"].font = std_font(NAVY, 8)
ws["A4"].alignment = left_align()
ws.row_dimensions[4].height = 28

if not v3_bench_a.empty:
    write_df_to_ws(ws, v3_bench_a, start_row=5, col_color_fn=color_by_nivel_riesgo)
elif not bench_a.empty:
    # Fallback: usar fact_benchmark filtrando factores
    fb = bench_a[bench_a["nivel_analisis"] == "factor"].copy() if "nivel_analisis" in bench_a.columns else bench_a
    write_df_to_ws(ws, fb, start_row=5, col_color_fn=color_by_nivel_riesgo)
else:
    ws.cell(row=5, column=1).value = "Sin datos"
print("  [OK] V3_Benchmarking")


# ─────────────────────────────────────────────────────────────────────────────
# V3 — RANKING ÁREAS
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("V3_Ranking_Areas")
write_title_block(ws,
    "V3 Paso 5 — Ranking Top 5 Áreas Críticas",
    "Script: 09_asis_gerencial.py | Áreas con mayor % Riesgo Alto+MuyAlto en fact_consolidado")

if not v3_rank_a.empty:
    write_df_to_ws(ws, v3_rank_a, start_row=4, col_color_fn=color_by_nivel_riesgo)
else:
    # Calcular ranking desde consolidado
    if not cons_a.empty and "area_departamento" in cons_a.columns and "nivel_riesgo" in cons_a.columns:
        df_rank = (
            cons_a[cons_a["nivel_analisis"] == "factor"]
            .groupby("area_departamento")
            .apply(lambda x: pd.Series({
                "n_total": len(x),
                "n_alto_muyalto": (x["nivel_riesgo"] >= 4).sum(),
                "pct_alto_muyalto": round(100 * (x["nivel_riesgo"] >= 4).mean(), 1)
            }))
            .reset_index()
            .sort_values("pct_alto_muyalto", ascending=False)
            .head(10)
        )
        write_df_to_ws(ws, df_rank, start_row=4, col_color_fn=color_by_nivel_riesgo)
    else:
        ws.cell(row=4, column=1).value = "Sin datos de áreas o nivel_riesgo"
print("  [OK] V3_Ranking_Areas")


# ─────────────────────────────────────────────────────────────────────────────
# ZZ_RESUMEN — RESUMEN EJECUTIVO
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("ZZ_RESUMEN")

ws.merge_cells("A1:D1")
ws["A1"].value = "RESUMEN EJECUTIVO DE AUDITORÍA — MentalPRO"
ws["A1"].fill = fill(NAVY)
ws["A1"].font = bold_font(WHITE, 16)
ws["A1"].alignment = center()
ws.row_dimensions[1].height = 40

ws.merge_cells("A2:D2")
ws["A2"].value = f"Empresa: {EMPRESA}  |  Fecha auditoría: {FECHA_AUDITORIA}  |  Pipeline: 3 Visualizadores (V1, V2, V3)"
ws["A2"].fill = fill(GOLD)
ws["A2"].font = bold_font(WHITE, 11)
ws["A2"].alignment = center()
ws.row_dimensions[2].height = 22

# Encabezados
for ci, h in enumerate(["CATEGORÍA", "INDICADOR", "VALOR"], 1):
    cell = ws.cell(row=3, column=ci)
    cell.value = h
    cell.fill = fill(NAVY)
    cell.font = bold_font(WHITE, 10)
    cell.alignment = center()
    cell.border = thin_border()
ws.row_dimensions[3].height = 20

cat_prev = None
for ri, (cat, ind, val) in enumerate(resumen_rows, 4):
    bg = GRAY_L if ri % 2 == 0 else WHITE
    cell_cat = ws.cell(row=ri, column=1)
    cell_ind = ws.cell(row=ri, column=2)
    cell_val = ws.cell(row=ri, column=3)

    # Agrupar visualmente por categoría
    if cat != cat_prev:
        cell_cat.value = cat
        cell_cat.font = bold_font(NAVY, 9)
        cat_prev = cat
    else:
        cell_cat.value = ""

    cell_ind.value = ind
    cell_val.value = val

    for cell in (cell_cat, cell_ind, cell_val):
        cell.fill = fill(bg)
        cell.font = std_font(size=9) if cell != cell_cat or cat == cat_prev else bold_font(NAVY, 9)
        cell.alignment = left_align()
        cell.border = thin_border()
    ws.row_dimensions[ri].height = 15

ws.column_dimensions["A"].width = 25
ws.column_dimensions["B"].width = 55
ws.column_dimensions["C"].width = 50
print("  [OK] ZZ_RESUMEN")


# ─────────────────────────────────────────────────────────────────────────────
# ZZ_ALERTAS — ALERTAS Y PUNTOS DE CHEQUEO
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("ZZ_ALERTAS")

ws.merge_cells("A1:F1")
ws["A1"].value = "ALERTAS Y PUNTOS DE CHEQUEO PARA AUDITORÍA — MentalPRO"
ws["A1"].fill = fill(NAVY)
ws["A1"].font = bold_font(WHITE, 14)
ws["A1"].alignment = center()
ws.row_dimensions[1].height = 36

ws.merge_cells("A2:F2")
ws["A2"].value = (
    f"Empresa: {EMPRESA}  |  "
    f"Total alertas: {len(alertas)}  |  "
    f"Críticas: {sum(1 for a in alertas if a['Nivel']=='CRITICO')}  |  "
    f"Advertencias: {sum(1 for a in alertas if a['Nivel']=='ADVERTENCIA')}  |  "
    f"Info: {sum(1 for a in alertas if a['Nivel']=='INFO')}"
)
ws["A2"].fill = fill(GOLD)
ws["A2"].font = bold_font(NAVY, 10)
ws["A2"].alignment = center()
ws.row_dimensions[2].height = 20

# Leyenda
leyenda = [
    ("CRITICO",      "EF4444", "Requiere acción inmediata antes de publicar resultados"),
    ("ADVERTENCIA",  "F97316", "Revisar y decidir si impacta la interpretación"),
    ("INFO",         "10B981", "Información de contexto, sin acción requerida"),
]
for i, (niv, color, desc) in enumerate(leyenda, 3):
    ws.merge_cells(f"A{i}:F{i}")
    ws[f"A{i}"].value = f"  {niv}: {desc}"
    ws[f"A{i}"].fill = fill(color)
    ws[f"A{i}"].font = bold_font(WHITE, 9)
    ws[f"A{i}"].alignment = left_align()
    ws.row_dimensions[i].height = 16

# Encabezados tabla alertas
headers = ["Nivel", "Código", "Descripción", "Valor encontrado", "Umbral / Referencia", "Acción recomendada"]
for ci, h in enumerate(headers, 1):
    cell = ws.cell(row=6, column=ci)
    cell.value = h
    cell.fill = fill(NAVY)
    cell.font = bold_font(WHITE, 10)
    cell.alignment = center()
    cell.border = thin_border()
ws.row_dimensions[6].height = 22

ALERTA_COLOR = {"CRITICO": "EF4444", "ADVERTENCIA": "F97316", "INFO": "A7F3D0"}

for ri, alerta in enumerate(alertas, 7):
    bg_nivel = ALERTA_COLOR.get(alerta["Nivel"], WHITE)
    ws.row_dimensions[ri].height = 16
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=ri, column=ci)
        cell.value = alerta.get(h, "")
        cell.border = thin_border()
        cell.alignment = left_align()
        if ci == 1:  # columna Nivel → colorear
            cell.fill = fill(bg_nivel)
            cell.font = bold_font(WHITE if alerta["Nivel"] != "INFO" else NAVY, 9)
        else:
            cell.fill = fill(GRAY_L if ri % 2 == 0 else WHITE)
            cell.font = std_font(size=9)

# Ancho columnas alertas
ws.column_dimensions["A"].width = 14
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 55
ws.column_dimensions["D"].width = 25
ws.column_dimensions["E"].width = 25
ws.column_dimensions["F"].width = 55

# Checklist adicional de puntos de chequeo manuales
checklist = [
    "",
    "═══════════════════════════════════════════════════════════════════════════════",
    "CHECKLIST MANUAL PARA AUDITOR — Verificación de calidad del pipeline ACVICOL",
    "═══════════════════════════════════════════════════════════════════════════════",
    "[ ] 1. Verificar que todos los trabajadores de ACVICOL tienen respuestas en los 6 instrumentos",
    "[ ] 2. Confirmar que la forma_intra (A/B) coincide con la categoría_cargo del trabajador",
    "[ ] 3. Validar que no hay PK duplicadas (cedula + forma_intra + id_pregunta) en fact_respuestas_clean",
    "[ ] 4. Revisar que el sector_economico fue homologado correctamente a un sector ENCST válido",
    "[ ] 5. Comparar manualmente 5-10 puntajes transformados vs tablas Res. 2764/2022 (Paso 9-14)",
    "[ ] 6. Verificar que ítems invertidos tienen valor_invertido ≠ valor_numerico en la hoja 02a_Paso2_Inversion",
    "[ ] 7. Confirmar que el referente sectorial usado en benchmarking corresponde al sector homologado",
    "[ ] 8. Revisar si hay dimensiones sin baremo asignado (nivel_riesgo nulo en hoja 02b_Baremos_Dim)",
    "[ ] 9. Verificar cobertura ausentismo — si <10% trabajadores tienen registro, señalar en informe",
    "[ ] 10. Confirmar que grupos demográficos <5 personas tienen datos marcados como 'Confidencial' (R8)",
    "[ ] 11. Validar que el cálculo de costos R10 usa los parámetros del config.yaml (no valores hardcodeados)",
    "[ ] 12. Revisar top 20 preguntas vs ENCST — confirmar que las comparables son las 39 documentadas",
    "[ ] 13. Verificar que protocolos V2 priorizados corresponden al sector de la empresa",
    "[ ] 14. Confirmar que no hay trabajadores con nivel_riesgo nulo en fact_scores_baremo (parcial ok si no aplica)",
    "[ ] 15. Validar que el ranking de áreas V3 coincide con los resultados de heatmap en V1 (S9/S10)",
]

row_cl = len(alertas) + 10
for i, linea in enumerate(checklist, row_cl):
    ws.merge_cells(f"A{i}:F{i}")
    cell = ws[f"A{i}"]
    cell.value = linea
    if "═══" in linea:
        cell.fill = fill(NAVY)
        cell.font = bold_font(WHITE, 9)
    elif linea.startswith("CHECKLIST"):
        cell.fill = fill(NAVY)
        cell.font = bold_font(WHITE, 10)
    elif linea.startswith("[ ]"):
        cell.fill = fill(GRAY_L if i % 2 == 0 else WHITE)
        cell.font = std_font(size=9)
    else:
        cell.fill = fill(WHITE)
        cell.font = std_font(size=9)
    cell.alignment = left_align()
    cell.border = thin_border()
    ws.row_dimensions[i].height = 14

print("  [OK] ZZ_ALERTAS")


# ══════════════════════════════════════════════════════════════════════════════
# GUARDAR ARCHIVO
# ══════════════════════════════════════════════════════════════════════════════
fecha_str = datetime.now().strftime("%Y%m%d_%H%M")
out_path = OUT / f"Auditoria_ETL_{EMPRESA}_{fecha_str}.xlsx"
wb.save(str(out_path))

print(f"\n{'='*60}")
print(f"  ARCHIVO GENERADO:")
print(f"  {out_path}")
print(f"{'='*60}")
print(f"  Hojas creadas: {len(wb.sheetnames)}")
print(f"  Alertas generadas: {len(alertas)}")
print(f"    — Críticas:    {sum(1 for a in alertas if a['Nivel']=='CRITICO')}")
print(f"    — Advertencias:{sum(1 for a in alertas if a['Nivel']=='ADVERTENCIA')}")
print(f"    — Info:        {sum(1 for a in alertas if a['Nivel']=='INFO')}")
print(f"{'='*60}\n")
