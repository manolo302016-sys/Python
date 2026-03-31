"""
06_vigilancia_epidemiologica.py
================================
ETL Visualizador 2 — Gestión de Salud Mental.

Paso 6: Insumo para el Programa de Vigilancia Epidemiológica (PVE).

  6.1  Evaluar los 11 indicadores priorizados para vigilancia contra el
       criterio de caso sospechoso de cada uno.
  6.2  Construir tabla resumen por empresa × indicador (n_casos, pct_casos).
  6.3  Listar nominativamente los trabajadores con sus criterios cumplidos,
       ordenados de mayor a menor número de criterios.
  6.4  Generar archivo de auditoría Excel (pasos 1-6).

Fuentes:
  data/processed/fact_gestion_04_niveles_indicadores.parquet  (principal)
  data/processed/dim_trabajador.parquet
  data/processed/ausentismo_12meses.parquet  (OPCIONAL → fallback: dim_ausentismo.parquet)

Outputs:
  data/processed/fact_gestion_06_vigilancia_resumen.parquet
  data/processed/fact_gestion_06_vigilancia_trabajadores.parquet
  data/processed/fact_gestion_06_vigilancia_ranking.parquet
  data/processed/audit_indecol_consultel_pasos1_6.xlsx
"""

import logging
import shutil
import sys
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)

# ══════════════════════════════════════════════════════════════════════════════
# Definición de los 11 indicadores de vigilancia epidemiológica
# Fuente: doc v2_gestion_saludmental.md — Paso 6
#
# Criterio aplicado sobre score_indicador (paso 3.1 — antes de inversión 3.2):
#   op="lte": score_indicador <= umbral  (indicadores directos: + score = mejor)
#   op="gte": score_indicador >= umbral  (indicadores de síntoma/riesgo: + score = más riesgo)
# ══════════════════════════════════════════════════════════════════════════════
INDICADORES_VIGILANCIA = [
    {
        "id": "VIG-01",
        "fuente": "intralaboral",
        "indicador": "Convivencia y respeto",
        "linea_gestion": "Convivencia y relacionamiento",
        "definicion": "Trabajadores con autoreporte de exposición a conductas de maltrato o irrespeto en el entorno laboral",
        "criterio_sospechoso": "score_indicador <= 0.45",
        "criterio_confirmado": "Trabajadores con queja de acoso laboral en comité de convivencia laboral",
        "mecanismo_confirmacion": "Reportes oficiales al CCL",
        "soporte_legal": "Res.2764/2022 / Ley 1010/2006",
        "enfoque": "Cuidado de la salud",
        "tipo": "score",
        "op": "lte",
        "umbral": 0.45,
    },
    {
        "id": "VIG-02",
        "fuente": "estres",
        "indicador": "Conductas de riesgo",
        "linea_gestion": "Comportamientos seguros",
        "definicion": "Trabajadores con autoreporte de comportamientos de riesgo como consumo de drogas, bebidas, accidentes",
        "criterio_sospechoso": "score_indicador >= 0.45",
        "criterio_confirmado": "Trabajadores con consumo problemático de alcohol y sustancias psicoactivas",
        "mecanismo_confirmacion": "Pruebas biológicas + pruebas de patrón de consumo",
        "soporte_legal": "Res.2764/2022 / Res. 1968/2025 / Dec 728/2025",
        "enfoque": "Cuidado de la salud",
        "tipo": "score",
        "op": "gte",
        "umbral": 0.45,
    },
    {
        "id": "VIG-03",
        "fuente": "ausentismo",
        "indicador": "Accidente de trabajo",
        "linea_gestion": None,
        "definicion": "Trabajadores con registro de accidentalidad laboral en los últimos 12 meses",
        "criterio_sospechoso": "> 1 evento de accidente laboral en tipo de ausentismo",
        "criterio_confirmado": "Trabajadores que han presentado accidentes de trabajo múltiples (más de 2 en 12 meses)",
        "mecanismo_confirmacion": "FURAT, formatos de investigación, reportes",
        "soporte_legal": "Res.2764/2022",
        "enfoque": "Cuidado de la salud",
        "tipo": "ausentismo_at",
    },
    {
        "id": "VIG-04",
        "fuente": "ausentismo",
        "indicador": "Ausentismo de interés psicosocial",
        "linea_gestion": None,
        "definicion": "Trabajadores que en los últimos 12 meses han presentado al menos 1 evento de ausentismo de interés psicosocial",
        "criterio_sospechoso": "1 o más eventos de ausentismo laboral con diagnóstico CIE de interés psicosocial",
        "criterio_confirmado": "Trabajadores con diagnósticos psicosociales CIE en los últimos 12 meses",
        "mecanismo_confirmacion": "Registros de ausentismo laboral",
        "soporte_legal": "Res.2764/2022",
        "enfoque": "Cuidado de la salud",
        "tipo": "ausentismo_cie",
    },
    {
        "id": "VIG-05",
        "fuente": "estres",
        "indicador": "Somatización y fatiga física",
        "linea_gestion": "Bienestar físico",
        "definicion": "Trabajadores con autoreporte de síntomas físicos asociados al estrés",
        "criterio_sospechoso": "score_indicador >= 0.45",
        "criterio_confirmado": "Trabajadores con evento vital estresante o alta exposición a factores de riesgo psicosocial",
        "mecanismo_confirmacion": "Valoración psicosocial individual",
        "soporte_legal": "Res.2764/2022 / Res. 2646/2008",
        "enfoque": "Cuidado de la salud",
        "tipo": "score",
        "op": "gte",
        "umbral": 0.45,
    },
    {
        "id": "VIG-06",
        "fuente": "estres",
        "indicador": "Desgaste emocional",
        "linea_gestion": "Bienestar emocional y trascendente",
        "definicion": "Trabajadores con autoreporte de síntomas emocionales asociados al estrés",
        "criterio_sospechoso": "score_indicador >= 0.45",
        "criterio_confirmado": "Trabajadores con evento vital estresante o alta exposición a factores de riesgo psicosocial",
        "mecanismo_confirmacion": "Valoración psicosocial individual",
        "soporte_legal": "Res.2764/2022 / Res. 2646/2008",
        "enfoque": "Cuidado de la salud",
        "tipo": "score",
        "op": "gte",
        "umbral": 0.45,
    },
    {
        "id": "VIG-07",
        "fuente": "estres",
        "indicador": "Desmotivación y desgaste laboral",
        "linea_gestion": "Motivación laboral",
        "definicion": "Trabajadores con autoreporte de desmotivación laboral, desgaste e intención de rotación",
        "criterio_sospechoso": "score_indicador >= 0.45",
        "criterio_confirmado": "Sin requisito legal directo",
        "mecanismo_confirmacion": "Valoración psicosocial individual",
        "soporte_legal": "Sin requisito legal directo",
        "enfoque": "Promoción del bienestar",
        "tipo": "score",
        "op": "gte",
        "umbral": 0.45,
    },
    {
        "id": "VIG-08",
        "fuente": "estres",
        "indicador": "Pérdida de sentido",
        "linea_gestion": "Bienestar emocional y trascendente",
        "definicion": "Trabajadores con autoreporte de pérdida de sentido de vida",
        "criterio_sospechoso": "score_indicador >= 0.45",
        "criterio_confirmado": "Sin requisito legal directo",
        "mecanismo_confirmacion": "Valoración psicosocial individual",
        "soporte_legal": "Sin requisito legal directo",
        "enfoque": "Promoción del bienestar",
        "tipo": "score",
        "op": "gte",
        "umbral": 0.45,
    },
    {
        "id": "VIG-09",
        "fuente": "intralaboral",
        "indicador": "Interferencia temporal",
        "linea_gestion": "Bienestar vida-trabajo",
        "definicion": "Trabajadores con autoreporte de alto conflicto del trabajo con la vida personal por interferencia de tiempo",
        "criterio_sospechoso": "score_indicador >= 0.45",
        "criterio_confirmado": "Sin requisito legal directo",
        "mecanismo_confirmacion": "Valoración psicosocial individual",
        "soporte_legal": "Sin requisito legal directo",
        "enfoque": "Promoción del bienestar",
        "tipo": "score",
        "op": "gte",
        "umbral": 0.45,
    },
    {
        "id": "VIG-10",
        "fuente": "extralaboral",
        "indicador": "Bienestar financiero",
        "linea_gestion": "Bienestar financiero",
        "definicion": "Trabajadores con autoreporte de estresores financieros y alto endeudamiento",
        "criterio_sospechoso": "score_indicador <= 0.45",
        "criterio_confirmado": "Sin requisito legal directo",
        "mecanismo_confirmacion": "Valoración psicosocial individual",
        "soporte_legal": "Sin requisito legal directo",
        "enfoque": "Promoción del bienestar",
        "tipo": "score",
        "op": "lte",
        "umbral": 0.45,
    },
    {
        "id": "VIG-11",
        "fuente": "extralaboral",
        "indicador": "Apoyo social",
        "linea_gestion": "Bienestar extralaboral",
        "definicion": "Trabajadores con autoreporte de baja calidad en red de apoyo social",
        "criterio_sospechoso": "score_indicador <= 0.45",
        "criterio_confirmado": "Sin requisito legal directo",
        "mecanismo_confirmacion": "Valoración psicosocial individual",
        "soporte_legal": "Sin requisito legal directo",
        "enfoque": "Promoción del bienestar",
        "tipo": "score",
        "op": "lte",
        "umbral": 0.45,
    },
]

# Semáforo VIG
EMPRESAS_AUDITORIA = ["INDECOL", "CONSULTEL"]


def _sanitizar_tipos(df: pd.DataFrame) -> pd.DataFrame:
    for col in df.select_dtypes(include=["object", "string"]).columns:
        df[col] = df[col].astype(str).where(df[col].notna(), None)
    return df


def _cargar_fuentes(root: Path) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame | None]:
    """Carga las fuentes principales. Ausentismo es opcional."""
    df_ind = pd.read_parquet(root / "data" / "processed" / "fact_gestion_04_niveles_indicadores.parquet")
    df_dim = pd.read_parquet(root / "data" / "processed" / "dim_trabajador.parquet")

    # Ausentismo: fallback chain
    path_aus_12 = root / "data" / "processed" / "ausentismo_12meses.parquet"
    path_aus_dim = root / "data" / "processed" / "dim_ausentismo.parquet"

    df_aus = None
    if path_aus_12.exists():
        df_aus = pd.read_parquet(path_aus_12)
        log.info("Ausentismo: cargado desde ausentismo_12meses.parquet (%d filas)", len(df_aus))
    elif path_aus_dim.exists():
        df_aus = pd.read_parquet(path_aus_dim)
        log.warning(
            "ausentismo_12meses.parquet no encontrado — usando dim_ausentismo.parquet (%d filas). "
            "Nota: puede contener datos de múltiples años.",
            len(df_aus),
        )
    else:
        log.warning("No se encontró ningún archivo de ausentismo. VIG-03 y VIG-04 se marcarán como 'Sin datos'.")

    return df_ind, df_dim, df_aus


def _evaluar_criterio_score(
    df_ind: pd.DataFrame,
    cfg: dict,
    df_dim: pd.DataFrame,
) -> pd.DataFrame:
    """
    Evalúa criterio score-based aplicado sobre score_indicador (paso 3.1),
    usando el operador original del doc V2:
      - op="gte": score_indicador >= umbral  (mayor score = mayor síntoma/riesgo)
      - op="lte": score_indicador <= umbral  (menor score = mayor riesgo)
    Para deduplicar A/B se conserva el peor caso según el operador:
      - gte: máximo score_indicador (más síntomas)
      - lte: mínimo score_indicador (peor gestión)
    """
    # Filtrar por indicador exacto + linea_gestion para distinguir homónimos
    mask = df_ind["indicador"] == cfg["indicador"]
    if cfg.get("linea_gestion"):
        mask = mask & (df_ind["linea_gestion"] == cfg["linea_gestion"])
    sub = df_ind[mask].copy()

    if sub.empty:
        log.warning("  %s: sin datos para indicador '%s'", cfg["id"], cfg["indicador"])
        return pd.DataFrame(columns=["cedula", "empresa", "vig_id"])

    # Deduplicar A/B: conservar el peor caso según dirección del operador
    if cfg["op"] == "gte":
        # Mayor score_indicador = peor (más síntomas) → conservar máximo
        sub = (
            sub.sort_values("score_indicador", ascending=False)
            .drop_duplicates(subset=["cedula", "empresa"], keep="first")
        )
        casos = sub[sub["score_indicador"] >= cfg["umbral"]][["cedula", "empresa"]].copy()
    else:  # lte
        # Menor score_indicador = peor → conservar mínimo
        sub = (
            sub.sort_values("score_indicador", ascending=True)
            .drop_duplicates(subset=["cedula", "empresa"], keep="first")
        )
        casos = sub[sub["score_indicador"] <= cfg["umbral"]][["cedula", "empresa"]].copy()
    casos["vig_id"] = cfg["id"]
    return casos.reset_index(drop=True)


def _evaluar_criterio_ausentismo_at(
    df_aus: pd.DataFrame,
    df_dim: pd.DataFrame,
    cfg: dict,
) -> pd.DataFrame:
    """
    VIG-03: Accidentes de trabajo — > 1 evento por trabajador.
    Retorna DataFrame con columnas: cedula, empresa, vig_id.
    """
    at = df_aus[df_aus["tipo_ausentismo"].str.strip().str.lower() == "accidente de trabajo"].copy()
    # Contar eventos reales (excluir prórrogas si existe la columna)
    if "es_prorroga" in at.columns:
        at = at[at["es_prorroga"].astype(str).str.strip().str.lower().isin(["no", "false", "0"])]
    conteo = at.groupby("cedula").size().reset_index(name="n_at")
    casos_ceds = conteo[conteo["n_at"] > 1]["cedula"].unique()

    # Obtener empresa vía dim_trabajador
    casos = (
        df_dim[df_dim["cedula"].isin(casos_ceds)][["cedula", "empresa"]]
        .drop_duplicates()
        .copy()
    )
    casos["vig_id"] = cfg["id"]
    return casos.reset_index(drop=True)


def _evaluar_criterio_ausentismo_cie(
    df_aus: pd.DataFrame,
    df_dim: pd.DataFrame,
    cfg: dict,
) -> pd.DataFrame:
    """
    VIG-04: Diagnósticos CIE de interés psicosocial.
    Prefijos: F (trastornos mentales), Z55-Z63 (factores sociales/laborales).
    Al menos 1 evento.
    """
    def _es_psicosocial(cie: str) -> bool:
        if not isinstance(cie, str):
            return False
        cie = cie.strip().upper()
        if cie.startswith("F"):
            return True
        # Z55–Z63
        for prefix in ["Z55", "Z56", "Z57", "Z58", "Z59", "Z60", "Z61", "Z62", "Z63"]:
            if cie.startswith(prefix):
                return True
        return False

    psico = df_aus[df_aus["diagnostico_CIE"].apply(_es_psicosocial)].copy()
    casos_ceds = psico["cedula"].unique()

    casos = (
        df_dim[df_dim["cedula"].isin(casos_ceds)][["cedula", "empresa"]]
        .drop_duplicates()
        .copy()
    )
    casos["vig_id"] = cfg["id"]
    return casos.reset_index(drop=True)


def _construir_tabla_trabajadores(
    criterios_dict: dict[str, pd.DataFrame],
    df_ind: pd.DataFrame,
    df_dim: pd.DataFrame,
) -> pd.DataFrame:
    """
    Consolida todos los criterios cumplidos por trabajador × empresa.
    Retorna: cedula, nombre_trabajador, empresa, area_departamento, categoria_cargo,
             criterios_cumplidos (lista separada por |), n_criterios.
    """
    if not any(len(v) > 0 for v in criterios_dict.values()):
        return pd.DataFrame()

    partes = []
    for vig_id, df_crit in criterios_dict.items():
        if len(df_crit) > 0:
            partes.append(df_crit[["cedula", "empresa", "vig_id"]])

    if not partes:
        return pd.DataFrame()

    long = pd.concat(partes, ignore_index=True)

    # Agrupar por cedula × empresa
    agrupado = (
        long.groupby(["cedula", "empresa"])["vig_id"]
        .apply(lambda x: "|".join(sorted(x)))
        .reset_index(name="criterios_cumplidos")
    )
    agrupado["n_criterios"] = agrupado["criterios_cumplidos"].str.count(r"\|") + 1

    # Agregar datos del trabajador (nombre, area, cargo)
    # Preferir nombre de fact_gestion si está disponible
    nombres = (
        df_ind[["cedula", "empresa", "nombre_trabajador"]]
        .drop_duplicates(subset=["cedula", "empresa"])
    )
    dem = (
        df_dim[["cedula", "empresa", "area_departamento", "categoria_cargo"]]
        .drop_duplicates(subset=["cedula", "empresa"])
    )

    tabla = agrupado.merge(nombres, on=["cedula", "empresa"], how="left")
    tabla = tabla.merge(dem, on=["cedula", "empresa"], how="left")

    return tabla.sort_values(["empresa", "n_criterios"], ascending=[True, False]).reset_index(drop=True)


def _construir_resumen_empresa(
    criterios_dict: dict[str, pd.DataFrame],
    df_ind: pd.DataFrame,
    df_dim: pd.DataFrame,
) -> pd.DataFrame:
    """
    Tabla resumen: empresa × indicador con n_casos, pct_casos, n_total, metadata.
    """
    # Total de trabajadores únicos por empresa (base del denominador)
    total_por_empresa = (
        df_dim.groupby("empresa")["cedula"]
        .nunique()
        .reset_index(name="n_total")
    )

    filas = []
    for cfg in INDICADORES_VIGILANCIA:
        df_crit = criterios_dict.get(cfg["id"], pd.DataFrame())

        if len(df_crit) == 0 and cfg["tipo"] in ("ausentismo_at", "ausentismo_cie"):
            # Sin datos de ausentismo
            for _, row in total_por_empresa.iterrows():
                filas.append({
                    "empresa": row["empresa"],
                    "vig_id": cfg["id"],
                    "indicador": cfg["indicador"],
                    "fuente": cfg["fuente"],
                    "definicion": cfg["definicion"],
                    "criterio_sospechoso": cfg["criterio_sospechoso"],
                    "criterio_confirmado": cfg["criterio_confirmado"],
                    "mecanismo_confirmacion": cfg["mecanismo_confirmacion"],
                    "soporte_legal": cfg["soporte_legal"],
                    "enfoque": cfg["enfoque"],
                    "n_total": row["n_total"],
                    "n_casos": 0,
                    "pct_casos": 0.0,
                    "estado_datos": "Sin datos ausentismo",
                })
        elif len(df_crit) == 0:
            for _, row in total_por_empresa.iterrows():
                filas.append({
                    "empresa": row["empresa"],
                    "vig_id": cfg["id"],
                    "indicador": cfg["indicador"],
                    "fuente": cfg["fuente"],
                    "definicion": cfg["definicion"],
                    "criterio_sospechoso": cfg["criterio_sospechoso"],
                    "criterio_confirmado": cfg["criterio_confirmado"],
                    "mecanismo_confirmacion": cfg["mecanismo_confirmacion"],
                    "soporte_legal": cfg["soporte_legal"],
                    "enfoque": cfg["enfoque"],
                    "n_total": 0,
                    "n_casos": 0,
                    "pct_casos": 0.0,
                    "estado_datos": "Sin indicador en datos",
                })
        else:
            por_empresa = df_crit.groupby("empresa").size().reset_index(name="n_casos")
            merged = total_por_empresa.merge(por_empresa, on="empresa", how="left").fillna({"n_casos": 0})
            merged["n_casos"] = merged["n_casos"].astype(int)
            merged["pct_casos"] = (merged["n_casos"] / merged["n_total"] * 100).round(1)
            merged["estado_datos"] = "OK"

            for _, row in merged.iterrows():
                filas.append({
                    "empresa": row["empresa"],
                    "vig_id": cfg["id"],
                    "indicador": cfg["indicador"],
                    "fuente": cfg["fuente"],
                    "definicion": cfg["definicion"],
                    "criterio_sospechoso": cfg["criterio_sospechoso"],
                    "criterio_confirmado": cfg["criterio_confirmado"],
                    "mecanismo_confirmacion": cfg["mecanismo_confirmacion"],
                    "soporte_legal": cfg["soporte_legal"],
                    "enfoque": cfg["enfoque"],
                    "n_total": row["n_total"],
                    "n_casos": row["n_casos"],
                    "pct_casos": row["pct_casos"],
                    "estado_datos": "OK",
                })

    df_resumen = pd.DataFrame(filas)
    return df_resumen.sort_values(["empresa", "vig_id"]).reset_index(drop=True)


def _construir_ranking(tabla_trabajadores: pd.DataFrame) -> pd.DataFrame:
    """
    Ranking de trabajadores por empresa, ordenado por n_criterios DESC.
    """
    if tabla_trabajadores.empty:
        return pd.DataFrame()

    cols = ["empresa", "cedula", "nombre_trabajador", "area_departamento",
            "categoria_cargo", "n_criterios", "criterios_cumplidos"]
    existing = [c for c in cols if c in tabla_trabajadores.columns]
    return (
        tabla_trabajadores[existing]
        .sort_values(["empresa", "n_criterios"], ascending=[True, False])
        .reset_index(drop=True)
    )


def _semaforo_vig(pct: float) -> str:
    if pct < 5:
        return "verde"
    elif pct < 15:
        return "amarillo"
    elif pct < 30:
        return "naranja"
    return "rojo"


def _color_semaforo(semaforo: str) -> str:
    return {
        "verde": "#10B981",
        "amarillo": "#F59E0B",
        "naranja": "#F97316",
        "rojo": "#EF4444",
    }.get(semaforo, "#FFFFFF")


def _generar_auditoria_excel(root: Path, df_resumen: pd.DataFrame, df_trabajadores: pd.DataFrame) -> None:
    """
    Copia audit_indecol_consultel_pasos1_5.xlsx y agrega 2 hojas de P6.
    """
    src = root / "data" / "processed" / "audit_indecol_consultel_pasos1_5.xlsx"
    dst = root / "data" / "processed" / "audit_indecol_consultel_pasos1_6.xlsx"

    if not src.exists():
        log.warning("No se encontró %s — omitiendo generación de Excel.", src.name)
        return

    shutil.copy2(src, dst)
    log.info("Copiado %s -> %s", src.name, dst.name)

    wb = openpyxl.load_workbook(dst)

    # ── Colores y estilos ──────────────────────────────────────────────────────
    NAVY   = "0A1628"
    GOLD   = "C9952A"
    WHITE  = "FFFFFF"
    LIGHT  = "F3F4F6"

    hdr_fill  = PatternFill("solid", fgColor=NAVY)
    hdr_font  = Font(bold=True, color=WHITE, size=10)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Hoja P6 Vigilancia Resumen ─────────────────────────────────────────────
    if "P6 Vigilancia Resumen" in wb.sheetnames:
        del wb["P6 Vigilancia Resumen"]
    ws_res = wb.create_sheet("P6 Vigilancia Resumen")

    empresas_audit = [e for e in EMPRESAS_AUDITORIA if e in df_resumen["empresa"].values]
    df_res_audit = df_resumen[df_resumen["empresa"].isin(empresas_audit)].copy()
    df_res_audit["semaforo"] = df_res_audit["pct_casos"].apply(_semaforo_vig)

    cols_res = ["empresa", "vig_id", "indicador", "fuente", "definicion",
                "criterio_sospechoso", "n_total", "n_casos", "pct_casos", "semaforo",
                "criterio_confirmado", "mecanismo_confirmacion", "soporte_legal", "enfoque", "estado_datos"]
    cols_res = [c for c in cols_res if c in df_res_audit.columns]
    headers_res = [c.replace("_", " ").title() for c in cols_res]

    for ci, h in enumerate(headers_res, 1):
        cell = ws_res.cell(row=1, column=ci, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = hdr_align
        cell.border = border

    for ri, row in enumerate(df_res_audit[cols_res].itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            cell = ws_res.cell(row=ri, column=ci, value=val)
            cell.alignment = cell_align
            cell.border = border
            # Colorear celda de semáforo
            col_name = cols_res[ci - 1]
            if col_name == "semaforo":
                hex_color = _color_semaforo(str(val)).lstrip("#")
                cell.fill = PatternFill("solid", fgColor=hex_color)
                cell.font = Font(bold=True, color=WHITE if str(val) in ("rojo", "naranja") else "000000")
            # Colorear fila si pct_casos > 30%
            if col_name == "pct_casos" and isinstance(val, (int, float)) and val > 30:
                cell.font = Font(bold=True, color="EF4444")

    # Autofit columnas (aproximado)
    col_widths_res = [12, 8, 28, 14, 45, 35, 10, 10, 12, 12, 45, 35, 35, 22, 18]
    for ci, w in enumerate(col_widths_res[:len(cols_res)], 1):
        ws_res.column_dimensions[get_column_letter(ci)].width = w
    ws_res.freeze_panes = "C2"

    # ── Hoja P6 Trabajadores Criterios ─────────────────────────────────────────
    if "P6 Trabajadores Criterios" in wb.sheetnames:
        del wb["P6 Trabajadores Criterios"]
    ws_trab = wb.create_sheet("P6 Trabajadores Criterios")

    df_trab_audit = df_trabajadores[df_trabajadores["empresa"].isin(empresas_audit)].copy() if not df_trabajadores.empty else pd.DataFrame()

    cols_trab = ["empresa", "cedula", "nombre_trabajador", "area_departamento",
                 "categoria_cargo", "n_criterios", "criterios_cumplidos"]
    cols_trab = [c for c in cols_trab if not df_trab_audit.empty and c in df_trab_audit.columns]
    headers_trab = [c.replace("_", " ").title() for c in cols_trab]

    for ci, h in enumerate(headers_trab, 1):
        cell = ws_trab.cell(row=1, column=ci, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = hdr_align
        cell.border = border

    if not df_trab_audit.empty:
        for ri, row in enumerate(df_trab_audit[cols_trab].itertuples(index=False), 2):
            for ci, val in enumerate(row, 1):
                cell = ws_trab.cell(row=ri, column=ci, value=val)
                cell.alignment = cell_align
                cell.border = border
                # Resaltar trabajadores con ≥3 criterios
                if cols_trab[ci - 1] == "n_criterios" and isinstance(val, int) and val >= 3:
                    cell.fill = PatternFill("solid", fgColor="EF4444")
                    cell.font = Font(bold=True, color=WHITE)

    col_widths_trab = [14, 12, 30, 25, 30, 14, 40]
    for ci, w in enumerate(col_widths_trab[:len(cols_trab)], 1):
        ws_trab.column_dimensions[get_column_letter(ci)].width = w
    ws_trab.freeze_panes = "C2"

    wb.save(dst)
    log.info("Excel guardado: %s (%d hojas)", dst.name, len(wb.sheetnames))


def main() -> None:
    log.info("=== 06_vigilancia_epidemiologica.py — Paso 6: Vigilancia Epidemiológica ===")

    df_ind, df_dim, df_aus = _cargar_fuentes(ROOT)
    log.info("Indicadores cargados: %d | Trabajadores dim: %d | Ausentismo: %s",
             len(df_ind), len(df_dim), f"{len(df_aus)} filas" if df_aus is not None else "Sin datos")

    # ── 6.1 Evaluar criterios ─────────────────────────────────────────────────
    log.info("\n--- 6.1 Evaluando criterios de caso sospechoso ---")
    criterios: dict[str, pd.DataFrame] = {}

    for cfg in INDICADORES_VIGILANCIA:
        if cfg["tipo"] == "score":
            df_crit = _evaluar_criterio_score(df_ind, cfg, df_dim)
        elif cfg["tipo"] == "ausentismo_at":
            if df_aus is not None:
                df_crit = _evaluar_criterio_ausentismo_at(df_aus, df_dim, cfg)
            else:
                df_crit = pd.DataFrame(columns=["cedula", "empresa", "vig_id"])
        elif cfg["tipo"] == "ausentismo_cie":
            if df_aus is not None:
                df_crit = _evaluar_criterio_ausentismo_cie(df_aus, df_dim, cfg)
            else:
                df_crit = pd.DataFrame(columns=["cedula", "empresa", "vig_id"])
        else:
            df_crit = pd.DataFrame(columns=["cedula", "empresa", "vig_id"])

        criterios[cfg["id"]] = df_crit
        log.info("  %s %-35s -> %d casos sospechosos",
                 cfg["id"], f"({cfg['indicador']})", len(df_crit))

    # ── 6.2 Resumen por empresa × indicador ────────────────────────────────────
    log.info("\n--- 6.2 Construyendo resumen por empresa ---")
    df_resumen = _construir_resumen_empresa(criterios, df_ind, df_dim)

    # ── 6.3 Tabla nominativa de trabajadores ───────────────────────────────────
    log.info("\n--- 6.3 Construyendo tabla nominativa de trabajadores ---")
    df_trabajadores = _construir_tabla_trabajadores(criterios, df_ind, df_dim)
    df_ranking      = _construir_ranking(df_trabajadores)

    # ── Log resumen por empresa ────────────────────────────────────────────────
    log.info("\n=== RESUMEN VIGILANCIA POR EMPRESA ===")
    for empresa in sorted(df_resumen["empresa"].unique()):
        sub_emp = df_resumen[df_resumen["empresa"] == empresa]
        total_trab = sub_emp["n_total"].iloc[0] if len(sub_emp) > 0 else 0

        if not df_trabajadores.empty and empresa in df_trabajadores["empresa"].values:
            n_con_criterio = len(df_trabajadores[df_trabajadores["empresa"] == empresa])
            n_multi = len(df_trabajadores[(df_trabajadores["empresa"] == empresa) & (df_trabajadores["n_criterios"] >= 3)])
        else:
            n_con_criterio = 0
            n_multi = 0

        pct_con_criterio = round(n_con_criterio / total_trab * 100, 1) if total_trab > 0 else 0
        log.info("\n  [%s] Total trabajadores: %d | Con >=1 criterio: %d (%.1f%%) | Con >=3 criterios: %d",
                 empresa, total_trab, n_con_criterio, pct_con_criterio, n_multi)

        # Top 5 indicadores por pct_casos
        top5 = sub_emp[sub_emp["estado_datos"] == "OK"].nlargest(5, "pct_casos")
        for _, r in top5.iterrows():
            sem = _semaforo_vig(r["pct_casos"])
            log.info("    %s %-35s %d casos (%.1f%%) [%s]",
                     r["vig_id"], r["indicador"], r["n_casos"], r["pct_casos"], sem)

        if not df_ranking.empty and empresa in df_ranking["empresa"].values:
            top3_trab = df_ranking[df_ranking["empresa"] == empresa].head(3)
            log.info("  Top 3 trabajadores con más criterios:")
            for _, r in top3_trab.iterrows():
                nombre = r.get("nombre_trabajador", r["cedula"])
                log.info("    %-30s  n_criterios=%d  [%s]",
                         nombre, r["n_criterios"], r["criterios_cumplidos"])

    # ── 6.4 Guardar parquets ──────────────────────────────────────────────────
    log.info("\n--- 6.4 Guardando outputs ---")
    out_res  = ROOT / "data" / "processed" / "fact_gestion_06_vigilancia_resumen.parquet"
    out_trab = ROOT / "data" / "processed" / "fact_gestion_06_vigilancia_trabajadores.parquet"
    out_rank = ROOT / "data" / "processed" / "fact_gestion_06_vigilancia_ranking.parquet"

    _sanitizar_tipos(df_resumen).to_parquet(out_res, index=False)
    log.info("  Guardado: %s (%d filas)", out_res.name, len(df_resumen))

    if not df_trabajadores.empty:
        _sanitizar_tipos(df_trabajadores).to_parquet(out_trab, index=False)
        log.info("  Guardado: %s (%d filas)", out_trab.name, len(df_trabajadores))
    else:
        pd.DataFrame().to_parquet(out_trab, index=False)
        log.info("  Guardado: %s (vacío — 0 casos)", out_trab.name)

    if not df_ranking.empty:
        _sanitizar_tipos(df_ranking).to_parquet(out_rank, index=False)
        log.info("  Guardado: %s (%d filas)", out_rank.name, len(df_ranking))
    else:
        pd.DataFrame().to_parquet(out_rank, index=False)
        log.info("  Guardado: %s (vacío)", out_rank.name)

    # ── Generar Excel de auditoría ────────────────────────────────────────────
    log.info("\n--- Generando Excel de auditoría pasos 1-6 ---")
    _generar_auditoria_excel(ROOT, df_resumen, df_ranking)

    log.info("\n=== Paso 6 COMPLETADO ===")


if __name__ == "__main__":
    main()
