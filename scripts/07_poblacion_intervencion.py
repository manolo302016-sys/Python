"""
07_poblacion_intervencion.py
============================
ETL Visualizador 2 — Gestión de Salud Mental.

Paso 7: Bases de datos operativas para selección de población a intervenir.

  7.1  Base protocolos × trabajadores:
       Para cada trabajador con score_linea < 0.65, identificar los protocolos
       que le aplican (vía linea_gestion) y su nivel de prioridad.
       Filtrable por protocolo → ver la población a intervenir.

  7.2  Base VIG criterios en formato largo (long):
       Una fila por trabajador × criterio VEP cumplido.
       Filtrable por VIG-ID → ver todos los trabajadores que cumplen ese criterio.

Inputs:
  data/processed/fact_gestion_04_niveles_lineas.parquet
  data/processed/dim_protocolos_lineas.parquet
  data/processed/dim_trabajador.parquet
  data/processed/fact_gestion_06_vigilancia_trabajadores.parquet
  data/processed/fact_gestion_06_vigilancia_resumen.parquet

Outputs:
  data/processed/fact_gestion_07_protocolos_poblacion.parquet
  data/processed/fact_gestion_07_vig_criterios_long.parquet
  data/processed/audit_indecol_consultel_pasos1_7.xlsx
"""

import logging
import shutil
import sys
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)

EMPRESAS_AUDITORIA = ["INDECOL", "CONSULTEL"]

# Colores por nivel de gestión (R11)
COLOR_NIVEL = {
    "Gestion prorrogable":         "10B981",
    "Gestion preventiva":          "6EE7B7",
    "Gestion de mejora selectiva": "F59E0B",
    "Intervencion correctiva":     "F97316",
    "Intervencion Urgente":        "EF4444",
}

# Niveles a incluir en la base de intervención (score < 0.65)
NIVELES_INTERVENCION = {"Gestion de mejora selectiva", "Intervencion correctiva", "Intervencion Urgente"}


def _sanitizar_tipos(df: pd.DataFrame) -> pd.DataFrame:
    for col in df.select_dtypes(include=["object", "string"]).columns:
        df[col] = df[col].astype(str).where(df[col].notna(), None)
    return df


# ══════════════════════════════════════════════════════════════════════════════
# 7.1 — Base protocolos × trabajadores
# ══════════════════════════════════════════════════════════════════════════════

def _construir_protocolos_poblacion(
    df_lin: pd.DataFrame,
    df_prot: pd.DataFrame,
    df_dim: pd.DataFrame,
) -> pd.DataFrame:
    """
    Por cada trabajador con score_linea < 0.65 (niveles de intervención):
      - Identifica los protocolos asociados a esa línea de gestión.
      - Deduplica forma A/B conservando la peor situación (menor score_linea).
      - Une datos demográficos (área, cargo).

    Resultado: una fila por empresa × cedula × protocolo × línea.
    Filtrable por protocolo_id → muestra quiénes intervenir y con qué urgencia.
    """
    # Filtrar solo niveles que requieren intervención
    df_int = df_lin[df_lin["nivel_gestion"].isin(NIVELES_INTERVENCION)].copy()

    # Deduplicar forma A/B: conservar el score más bajo (peor) por cedula × empresa × linea
    df_int = (
        df_int.sort_values("score_linea", ascending=True)
        .drop_duplicates(subset=["cedula", "empresa", "linea_gestion"], keep="first")
    )

    # Unir protocolos por línea de gestión
    df_merged = df_int.merge(
        df_prot[["linea_gestion", "protocolo_id", "protocolo_nombre", "objetivo", "resultado_esperado"]],
        on="linea_gestion",
        how="left",
    )

    # Unir datos demográficos del trabajador
    cols_dim = ["cedula", "empresa", "area_departamento", "categoria_cargo", "tipo_cargo"]
    df_merged = df_merged.merge(
        df_dim[cols_dim].drop_duplicates(subset=["cedula", "empresa"]),
        on=["cedula", "empresa"],
        how="left",
    )

    # Columnas finales ordenadas
    cols_out = [
        "empresa", "protocolo_id", "protocolo_nombre",
        "cedula", "nombre_trabajador", "area_departamento", "categoria_cargo", "tipo_cargo",
        "linea_gestion", "eje_gestion", "score_linea", "nivel_gestion", "orden_nivel",
        "objetivo", "resultado_esperado",
    ]
    cols_out = [c for c in cols_out if c in df_merged.columns]

    df_out = (
        df_merged[cols_out]
        .sort_values(
            ["empresa", "protocolo_id", "orden_nivel", "score_linea"],
            ascending=[True, True, False, True],  # orden_nivel DESC (urgente=5 primero), score ASC
        )
        .reset_index(drop=True)
    )

    return df_out


# ══════════════════════════════════════════════════════════════════════════════
# 7.1b — Resumen trabajadores × protocolos
# ══════════════════════════════════════════════════════════════════════════════

def _construir_resumen_trabajadores(df_prot_pob: pd.DataFrame) -> pd.DataFrame:
    """
    Por cada trabajador: conteo de protocolos únicos en cada nivel de prioridad.
    Deduplica protocolo × trabajador conservando el peor nivel (mayor orden_nivel).

    Columnas resultado:
      empresa, cedula, nombre_trabajador, area_departamento, categoria_cargo, tipo_cargo,
      n_protocolos_total, n_urgente, n_correctiva, n_mejora_selectiva,
      nivel_max, protocolos_urgente, protocolos_correctiva
    """
    # Deduplicar: un protocolo puede venir de varias líneas → conservar la peor
    df_base = (
        df_prot_pob.sort_values("orden_nivel", ascending=False)
        .drop_duplicates(subset=["cedula", "empresa", "protocolo_id"], keep="first")
    )

    # Conteos por nivel
    def _conteo_nivel(grupo, nivel):
        return (grupo["nivel_gestion"] == nivel).sum()

    def _lista_prots(grupo, nivel):
        ids = sorted(grupo.loc[grupo["nivel_gestion"] == nivel, "protocolo_id"].tolist())
        return "|".join(ids) if ids else ""

    grp = df_base.groupby(["empresa", "cedula", "nombre_trabajador",
                           "area_departamento", "categoria_cargo", "tipo_cargo"],
                          dropna=False)

    rows = []
    for keys, g in grp:
        empresa, cedula, nombre, area, cargo, tipo = keys
        rows.append({
            "empresa":             empresa,
            "cedula":              cedula,
            "nombre_trabajador":   nombre,
            "area_departamento":   area,
            "categoria_cargo":     cargo,
            "tipo_cargo":          tipo,
            "n_protocolos_total":  len(g),
            "n_urgente":           _conteo_nivel(g, "Intervencion Urgente"),
            "n_correctiva":        _conteo_nivel(g, "Intervencion correctiva"),
            "n_mejora_selectiva":  _conteo_nivel(g, "Gestion de mejora selectiva"),
            "nivel_max":           g["nivel_gestion"].iloc[0],   # ya ordenado por orden DESC
            "protocolos_urgente":  _lista_prots(g, "Intervencion Urgente"),
            "protocolos_correctiva": _lista_prots(g, "Intervencion correctiva"),
        })

    df_res = pd.DataFrame(rows)
    df_res["n_criticos"] = df_res["n_urgente"] + df_res["n_correctiva"]

    # Reordenar columnas
    col_order = [
        "empresa", "cedula", "nombre_trabajador", "area_departamento",
        "categoria_cargo", "tipo_cargo",
        "n_protocolos_total", "n_urgente", "n_correctiva", "n_mejora_selectiva", "n_criticos",
        "nivel_max", "protocolos_urgente", "protocolos_correctiva",
    ]
    df_res = df_res[[c for c in col_order if c in df_res.columns]]

    return df_res.sort_values(
        ["empresa", "n_urgente", "n_correctiva", "n_mejora_selectiva"],
        ascending=[True, False, False, False],
    ).reset_index(drop=True)


# ══════════════════════════════════════════════════════════════════════════════
# 7.2 — VIG criterios en formato largo
# ══════════════════════════════════════════════════════════════════════════════

def _construir_vig_criterios_long(
    df_trab: pd.DataFrame,
    df_resumen: pd.DataFrame,
) -> pd.DataFrame:
    """
    Explota la columna criterios_cumplidos (pipe-separated) para obtener
    una fila por trabajador × criterio VEP.
    Une metadatos del indicador desde df_resumen.
    Resultado filtrable por vig_id → ver todos los trabajadores que lo cumplen.
    """
    if df_trab.empty:
        return pd.DataFrame()

    # Explotar criterios_cumplidos
    df_exp = df_trab.copy()
    df_exp["vig_id"] = df_exp["criterios_cumplidos"].str.split("|")
    df_exp = df_exp.explode("vig_id").reset_index(drop=True)
    df_exp["vig_id"] = df_exp["vig_id"].str.strip()

    # Metadatos por VIG-ID desde el resumen (una empresa cualquiera sirve)
    meta_cols = ["vig_id", "indicador", "fuente", "criterio_sospechoso", "soporte_legal", "enfoque"]
    meta_cols = [c for c in meta_cols if c in df_resumen.columns]
    meta = (
        df_resumen[meta_cols]
        .drop_duplicates(subset=["vig_id"])
    )

    df_long = df_exp.merge(meta, on="vig_id", how="left")

    # Columnas finales
    cols_out = [
        "empresa", "vig_id", "indicador", "fuente", "criterio_sospechoso",
        "soporte_legal", "enfoque",
        "cedula", "nombre_trabajador", "area_departamento", "categoria_cargo",
        "n_criterios",
    ]
    cols_out = [c for c in cols_out if c in df_long.columns]

    df_long = (
        df_long[cols_out]
        .sort_values(
            ["empresa", "vig_id", "n_criterios"],
            ascending=[True, True, False],  # trabajadores con más criterios primero
        )
        .reset_index(drop=True)
    )

    return df_long


# ══════════════════════════════════════════════════════════════════════════════
# Excel — agregar hojas P7
# ══════════════════════════════════════════════════════════════════════════════

def _estilo_hdr(ws, n_cols: int) -> None:
    """Aplica header navy a la fila 1."""
    hdr_fill  = PatternFill("solid", fgColor="0A1628")
    hdr_font  = Font(bold=True, color="FFFFFF", size=10)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=ci)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = hdr_align
        cell.border = border


def _escribir_df_en_hoja(
    ws,
    df: pd.DataFrame,
    col_color: str | None = None,       # columna cuyo valor determina el color de fila
    color_map: dict | None = None,      # valor → hex color
    col_widths: list[int] | None = None,
    freeze: str = "C2",
) -> None:
    """Escribe df en ws con header y formato."""
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell_align = Alignment(vertical="center", wrap_text=True)
    WHITE = "FFFFFF"

    headers = list(df.columns)
    _estilo_hdr(ws, len(headers))
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci).value = h.replace("_", " ").title()

    col_color_idx = headers.index(col_color) + 1 if col_color and col_color in headers else None

    for ri, row_vals in enumerate(df.itertuples(index=False), 2):
        # Determinar color de fila
        row_hex = None
        if col_color_idx and color_map:
            val = row_vals[col_color_idx - 1]
            row_hex = color_map.get(str(val))

        for ci, val in enumerate(row_vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = cell_align
            cell.border = border
            if row_hex:
                cell.fill = PatternFill("solid", fgColor=row_hex)
                # Texto blanco en niveles oscuros
                if row_hex in ("F97316", "EF4444"):
                    cell.font = Font(color=WHITE)

    # Anchos de columna
    if col_widths:
        for ci, w in enumerate(col_widths[:len(headers)], 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

    if freeze:
        ws.freeze_panes = freeze


def _generar_excel(
    root: Path,
    df_prot_pob: pd.DataFrame,
    df_vig_long: pd.DataFrame,
    df_resumen_trab: pd.DataFrame,
) -> None:
    src = root / "data" / "processed" / "audit_indecol_consultel_pasos1_6.xlsx"
    dst = root / "data" / "processed" / "audit_indecol_consultel_pasos1_7.xlsx"

    if not src.exists():
        log.warning("No se encontró %s — omitiendo Excel.", src.name)
        return

    shutil.copy2(src, dst)
    wb = openpyxl.load_workbook(dst)

    # ── Hoja P7 Protocolos Poblacion ──────────────────────────────────────────
    sheet_name = "P7 Protocolos Poblacion"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws_prot = wb.create_sheet(sheet_name)

    df_p = df_prot_pob[df_prot_pob["empresa"].isin(EMPRESAS_AUDITORIA)].copy()
    # Redondear score para legibilidad
    if "score_linea" in df_p.columns:
        df_p["score_linea"] = df_p["score_linea"].round(3)

    col_widths_prot = [14, 10, 40, 12, 30, 25, 28, 28, 35, 30, 12, 28, 12, 50, 45]
    _escribir_df_en_hoja(
        ws_prot,
        df_p,
        col_color="nivel_gestion",
        color_map=COLOR_NIVEL,
        col_widths=col_widths_prot,
        freeze="D2",
    )
    # Activar autofiltro
    ws_prot.auto_filter.ref = ws_prot.dimensions

    log.info("Hoja '%s': %d filas (INDECOL+CONSULTEL)", sheet_name, len(df_p))

    # ── Hoja P7 VIG Criterios Long ─────────────────────────────────────────────
    sheet_name2 = "P7 VIG Criterios Long"
    if sheet_name2 in wb.sheetnames:
        del wb[sheet_name2]
    ws_vig = wb.create_sheet(sheet_name2)

    df_v = df_vig_long[df_vig_long["empresa"].isin(EMPRESAS_AUDITORIA)].copy() if not df_vig_long.empty else pd.DataFrame()

    # Color por n_criterios: rojo si >=3, naranja si ==2, amarillo si ==1
    def _color_n(n):
        try:
            n = int(n)
        except (ValueError, TypeError):
            return None
        if n >= 3:
            return "EF4444"
        if n == 2:
            return "F97316"
        return "F59E0B"

    col_widths_vig = [14, 8, 30, 14, 30, 30, 22, 12, 30, 25, 28, 10]
    _escribir_df_en_hoja(
        ws_vig,
        df_v,
        col_color=None,
        color_map=None,
        col_widths=col_widths_vig,
        freeze="C2",
    )

    # Colorear columna n_criterios manualmente
    if not df_v.empty and "n_criterios" in df_v.columns:
        n_col_idx = list(df_v.columns).index("n_criterios") + 1
        WHITE = "FFFFFF"
        for ri in range(2, len(df_v) + 2):
            cell = ws_vig.cell(row=ri, column=n_col_idx)
            hex_c = _color_n(cell.value)
            if hex_c:
                cell.fill = PatternFill("solid", fgColor=hex_c)
                if hex_c in ("EF4444", "F97316"):
                    cell.font = Font(bold=True, color=WHITE)

    ws_vig.auto_filter.ref = ws_vig.dimensions
    log.info("Hoja '%s': %d filas (INDECOL+CONSULTEL)", sheet_name2, len(df_v))

    # ── Hoja P7 Resumen Trabajadores ──────────────────────────────────────────
    sheet_name3 = "P7 Resumen Trabajadores"
    if sheet_name3 in wb.sheetnames:
        del wb[sheet_name3]
    ws_res = wb.create_sheet(sheet_name3)

    df_r = df_resumen_trab[df_resumen_trab["empresa"].isin(EMPRESAS_AUDITORIA)].copy()

    col_widths_res = [14, 12, 30, 28, 28, 20, 14, 12, 14, 18, 10, 28, 22, 22]
    _escribir_df_en_hoja(
        ws_res,
        df_r,
        col_color="nivel_max",
        color_map=COLOR_NIVEL,
        col_widths=col_widths_res,
        freeze="C2",
    )

    # Resaltar columnas n_urgente y n_correctiva en rojo/naranja cuando > 0
    cols_r = list(df_r.columns)
    idx_urg = cols_r.index("n_urgente") + 1 if "n_urgente" in cols_r else None
    idx_cor = cols_r.index("n_correctiva") + 1 if "n_correctiva" in cols_r else None
    WHITE = "FFFFFF"
    for ri in range(2, len(df_r) + 2):
        if idx_urg:
            cell_u = ws_res.cell(row=ri, column=idx_urg)
            try:
                if int(cell_u.value or 0) > 0:
                    cell_u.fill = PatternFill("solid", fgColor="EF4444")
                    cell_u.font = Font(bold=True, color=WHITE)
            except (ValueError, TypeError):
                pass
        if idx_cor:
            cell_c = ws_res.cell(row=ri, column=idx_cor)
            try:
                if int(cell_c.value or 0) > 0:
                    cell_c.fill = PatternFill("solid", fgColor="F97316")
                    cell_c.font = Font(bold=True, color=WHITE)
            except (ValueError, TypeError):
                pass

    ws_res.auto_filter.ref = ws_res.dimensions
    log.info("Hoja '%s': %d filas (INDECOL+CONSULTEL)", sheet_name3, len(df_r))

    wb.save(dst)
    log.info("Excel guardado: %s (%d hojas)", dst.name, len(wb.sheetnames))


# ══════════════════════════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════════════════════════

def main() -> None:
    log.info("=== 07_poblacion_intervencion.py — Paso 7: Bases de datos operativas ===")

    # Cargar fuentes
    df_lin   = pd.read_parquet(ROOT / "data" / "processed" / "fact_gestion_04_niveles_lineas.parquet")
    df_prot  = pd.read_parquet(ROOT / "data" / "processed" / "dim_protocolos_lineas.parquet")
    df_dim   = pd.read_parquet(ROOT / "data" / "processed" / "dim_trabajador.parquet")
    df_trab  = pd.read_parquet(ROOT / "data" / "processed" / "fact_gestion_06_vigilancia_trabajadores.parquet")
    df_res   = pd.read_parquet(ROOT / "data" / "processed" / "fact_gestion_06_vigilancia_resumen.parquet")

    log.info("Lineas: %d | Protocolos dim: %d | Trabajadores dim: %d | VIG trab: %d",
             len(df_lin), len(df_prot), len(df_dim), len(df_trab))

    # ── 7.1 Base protocolos × trabajadores ────────────────────────────────────
    log.info("\n--- 7.1 Construyendo base protocolos x poblacion ---")
    df_prot_pob = _construir_protocolos_poblacion(df_lin, df_prot, df_dim)
    log.info("Resultado: %d filas | %d trabajadores únicos | %d protocolos únicos",
             len(df_prot_pob),
             df_prot_pob["cedula"].nunique(),
             df_prot_pob["protocolo_id"].nunique())

    # Resumen por empresa y nivel
    log.info("\n=== RESUMEN PROTOCOLOS x EMPRESA ===")
    for empresa in sorted(df_prot_pob["empresa"].unique()):
        sub = df_prot_pob[df_prot_pob["empresa"] == empresa]
        n_trab = sub["cedula"].nunique()
        n_prot = sub["protocolo_id"].nunique()
        dist = sub.drop_duplicates(subset=["cedula","linea_gestion"])["nivel_gestion"].value_counts().to_dict()
        log.info("  [%s] %d trabajadores | %d protocolos | Niveles: %s", empresa, n_trab, n_prot, dist)
        # Top 3 protocolos con más trabajadores urgentes/correctivos
        urgentes = sub[sub["nivel_gestion"].isin({"Intervencion correctiva","Intervencion Urgente"})]
        top3 = urgentes.groupby("protocolo_id")["cedula"].nunique().nlargest(3)
        for prot_id, n in top3.items():
            nombre = sub[sub["protocolo_id"]==prot_id]["protocolo_nombre"].iloc[0] if len(sub[sub["protocolo_id"]==prot_id]) > 0 else ""
            log.info("    %s %-45s -> %d en correctiva/urgente", prot_id, nombre[:45], n)

    # ── 7.2 VIG criterios formato largo ───────────────────────────────────────
    log.info("\n--- 7.2 Construyendo base VIG criterios (formato largo) ---")
    df_vig_long = _construir_vig_criterios_long(df_trab, df_res)
    log.info("Resultado: %d filas | %d VIG criterios únicos | %d trabajadores únicos",
             len(df_vig_long),
             df_vig_long["vig_id"].nunique() if not df_vig_long.empty else 0,
             df_vig_long["cedula"].nunique() if not df_vig_long.empty else 0)

    if not df_vig_long.empty:
        log.info("\n=== RESUMEN VIG CRITERIOS (total dataset) ===")
        resumen_vig = (
            df_vig_long.groupby("vig_id")
            .agg(n_trab=("cedula","nunique"), indicador=("indicador","first"))
            .sort_values("n_trab", ascending=False)
        )
        for vig_id, row in resumen_vig.iterrows():
            log.info("  %s %-35s -> %d trabajadores", vig_id, row["indicador"], row["n_trab"])

    # ── 7.1b Resumen trabajadores × protocolos ────────────────────────────────
    log.info("\n--- 7.1b Construyendo resumen trabajadores x protocolos ---")
    df_resumen_trab = _construir_resumen_trabajadores(df_prot_pob)
    log.info("Resultado: %d trabajadores únicos", len(df_resumen_trab))

    if not df_resumen_trab.empty:
        log.info("\n=== RESUMEN GLOBAL POR EMPRESA ===")
        for empresa in sorted(df_resumen_trab["empresa"].unique()):
            sub = df_resumen_trab[df_resumen_trab["empresa"] == empresa]
            log.info(
                "  [%s] %d trab | Urgente: %d | Correctiva: %d | Mejora selectiva: %d | Criticos (U+C): %d",
                empresa,
                len(sub),
                (sub["n_urgente"] > 0).sum(),
                (sub["n_correctiva"] > 0).sum(),
                (sub["n_mejora_selectiva"] > 0).sum(),
                (sub["n_criticos"] > 0).sum(),
            )

    # ── Guardar parquets ───────────────────────────────────────────────────────
    log.info("\n--- Guardando outputs ---")
    out_pp  = ROOT / "data" / "processed" / "fact_gestion_07_protocolos_poblacion.parquet"
    out_vl  = ROOT / "data" / "processed" / "fact_gestion_07_vig_criterios_long.parquet"
    out_res = ROOT / "data" / "processed" / "fact_gestion_07_resumen_trabajadores.parquet"

    _sanitizar_tipos(df_prot_pob).to_parquet(out_pp, index=False)
    log.info("  %s (%d filas)", out_pp.name, len(df_prot_pob))

    if not df_vig_long.empty:
        _sanitizar_tipos(df_vig_long).to_parquet(out_vl, index=False)
        log.info("  %s (%d filas)", out_vl.name, len(df_vig_long))
    else:
        pd.DataFrame().to_parquet(out_vl, index=False)
        log.info("  %s (vacio)", out_vl.name)

    _sanitizar_tipos(df_resumen_trab).to_parquet(out_res, index=False)
    log.info("  %s (%d filas)", out_res.name, len(df_resumen_trab))

    # ── Generar Excel ──────────────────────────────────────────────────────────
    log.info("\n--- Generando Excel audit pasos 1-7 ---")
    _generar_excel(ROOT, df_prot_pob, df_vig_long, df_resumen_trab)

    log.info("\n=== Paso 7 COMPLETADO ===")


if __name__ == "__main__":
    main()
